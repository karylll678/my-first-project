# -*- coding: utf-8 -*-
"""
券商数据处理程序稳定版（保留原核心逻辑，仅新增华泰/广发/中信等字段兼容）
功能：
  1. 将各券商原始交易流水数据统一格式化，并支持合并、归纳汇总
  2. 基于处理后的交易流水，优先使用通达信收盘价，股票池表兜底，模拟持仓变动，计算周度净值并生成图表
支持券商：广发、国君、华泰、建投、万联、爱建、中信、长江、申万
"""

# 稳定版说明：
# 1. 保留原有证券代码逻辑：港股5位，A股/ETF/北交所6位，不再统一zfill(6)。
# 2. 仅新增券商字段映射兼容：华泰、广发、中信等。
# 3. 不联网取价：通达信价格优先，股票池价格兜底。
# 4. 年末12月31日强制加入净值计算。

import os
import re
import json
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date, timedelta
from collections import defaultdict

import pandas as pd
import openpyxl

os.environ.setdefault(
    'MPLCONFIGDIR',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), '.matplotlib')
)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import font_manager


# ══════════════════════════════════════════════
# 配色方案（自定义现代风格）
# ══════════════════════════════════════════════
COLORS = {
    'bg':              '#F0F2F5',
    'card_bg':         '#FFFFFF',
    'accent':          '#2B6CB0',
    'accent_light':    '#EBF4FF',
    'accent_hover':    '#3182CE',
    'action_bg':       '#F6E05E',
    'action_fg':       '#744210',
    'action_hover':    '#ECC94B',
    'secondary_bg':    '#EDF2F7',
    'secondary_fg':    '#2D3748',
    'secondary_hover': '#E2E8F0',
    'open_bg':         '#C6F6D5',
    'open_fg':         '#22543D',
    'open_hover':      '#9AE6B4',
    'label_fg':        '#2D3748',
    'entry_fg':        '#4A5568',
    'entry_bg':        '#FAFAFA',
    'frame_border':    '#CBD5E0',
    'title_bg':        '#2B6CB0',
    'title_fg':        '#FFFFFF',
    'text_bg':         '#FAFAFA',
    'text_fg':         '#1A202C',
    'select_bg':       '#BEE3F8',
    'step_number':     '#2B6CB0',
    'step_label':      '#2D3748',
}

FONTS = {
    'title':       ('Microsoft YaHei UI', 14, 'bold'),
    'subtitle':    ('Microsoft YaHei UI', 11, 'bold'),
    'step':        ('Microsoft YaHei UI', 10, 'bold'),
    'button':      ('Microsoft YaHei UI', 9),
    'button_bold': ('Microsoft YaHei UI', 9, 'bold'),
    'entry':       ('Consolas', 9),
    'label':       ('Microsoft YaHei UI', 9),
    'text':        ('Consolas', 9),
    'hint':        ('Microsoft YaHei UI', 8),
}


# ─────────────────────────────────────────────
# 各券商原始数据的列名映射配置
# ─────────────────────────────────────────────
BROKER_CONFIG = {
    '广发': ('业务日期',   '业务标志名称', '证券代码', '证券名称', '成交数量', '成交价格', '清算金额', 0),
    '万联': ('日期',       '业务名称',     '证券代码', '证券名称', '成交数量', '成交均价', '发生金额', 0),
    '国君': ('交收日期',   '交易类别',     '证券代码', '证券名称', '成交数量', '成交价格', '成交金额', 0),
    # 华泰对账单字段：成交日期/业务名称/证券代码/证券名称/成交数量/成交均价/发生金额
    # 兼容只有“对账单”、没有“交割单”的华泰文件。
    '华泰': ('成交日期',   '业务名称',     '证券代码', '证券名称', '成交数量', '成交均价', '发生金额', 0),
    '建投': ('成交日期',   '业务名称',     '证券代码', '证券名称', '成交数量', '成交价格', '成交金额', 0),
    '爱建': ('成交日期',   '业务名称',     '证券代码', '证券名称', '成交数量', '成交价格', '发生金额', 0),
    '中信': ('发生日期',   '业务名称',     '证券代码', '证券名称', '成交数量', '成交价格', '成交金额', 0),
    '长江': ('清算日期',   '业务名称',     '证券代码', '证券名称', '成交数量', '成交价格', '发生金额', 0),
    '申万': ('交收日期',   '业务名称',     '证券代码', '证券名称', '成交数量', '成交均价', '成交金额', 0),
}

STANDARD_COLS = ['账户名称', '成交日期', '业务名称', '证券代码', '证券名称', '成交数量', '成交价格', '发生金额']
OPTIONAL_TXN_COLS = ['备注', '摘要', '操作', '交易市场', '市场', '资金本次余额', '资金余额', '资金余额(人民币)', '股份余额', '发生金额', '清算金额', '资金发生数', '成交金额']

# ─────────────────────────────────────────────
# 全局状态
# ─────────────────────────────────────────────
state = {
    'path_src':    '',
    'path_fmt':    '',
    'path_merge':  '',
    'path_induct': '',
    'book':        None,
    'df_merged':   None,
    'df_inducted': None,
    'save_path':   '',
    'save_path1':  '',
    'save_path2':  '',
}

# ── 净值计算相关路径 ──
# 默认从本程序所在目录读取/写入文件。当前稳定版放在 apython 文件夹时，
# 下面这些文件都会自动指向 /Users/kary/Desktop/apython。
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
LOGIC_FILE  = os.path.join(BASE_DIR, '交易流水业务变化逻辑.xlsx')
INIT_FILE   = os.path.join(BASE_DIR, '万联正涵250101_251231.xlsx')
TDX_PRICE_FILE = os.path.join(BASE_DIR, '通达信提取收盘价结果.xlsx')
PRICE_FILE     = os.path.join(BASE_DIR, '股票池2025年收盘价.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, '万联2025净值.xlsx')
CHART_FILE  = os.path.join(BASE_DIR, '万联2025净值图.png')


def configure_matplotlib_chinese_font():
    """选择当前电脑可用的中文字体，避免净值图中文显示成方块。"""
    preferred_fonts = [
        'PingFang SC', 'Hiragino Sans GB', 'Arial Unicode MS',
        'Songti SC', 'STHeiti', 'Heiti TC', 'Microsoft YaHei',
        'SimHei', 'Noto Sans CJK SC'
    ]
    available_fonts = {f.name for f in font_manager.fontManager.ttflist}
    for font_name in preferred_fonts:
        if font_name in available_fonts:
            plt.rcParams['font.sans-serif'] = [font_name, 'DejaVu Sans']
            break
    else:
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False


# ══════════════════════════════════════════════
# 一、文件路径选择函数
# ══════════════════════════════════════════════

def select_src_file():
    path = filedialog.askopenfilename()
    if path:
        state['path_src'] = path
        var_path_src.set(path)


def select_fmt_file():
    path = filedialog.askopenfilename()
    if not path:
        return
    if '数据表格样式' in path:
        state['path_fmt'] = path
        var_path_fmt.set(path)
    else:
        messagebox.showerror('文件错误', '请确认所选文件名称为"数据表格样式"')


def select_merge_dir():
    path = filedialog.askdirectory()
    if path:
        state['path_merge'] = path
        var_path_merge.set(path)


def select_induct_file():
    path = filedialog.askopenfilename()
    if path:
        state['path_induct'] = path
        var_path_induct.set(path)


# ══════════════════════════════════════════════
# 二、文件保存函数
# ══════════════════════════════════════════════

def save_book():
    path = filedialog.asksaveasfilename()
    if not path:
        return
    full_path = path + '.xlsx'
    state['book'].to_excel(full_path, sheet_name='合并', index=False)
    state['save_path'] = full_path
    var_save.set(full_path)


def save_inducted():
    path = filedialog.asksaveasfilename()
    if not path:
        return
    full_path = path + '.xlsx'
    state['df_inducted'].to_excel(full_path, sheet_name='归纳', index=False)
    state['save_path1'] = full_path
    var_save1.set(full_path)


def save_merged():
    path = filedialog.asksaveasfilename()
    if not path:
        return
    full_path = path + '.xlsx'
    state['df_merged'].to_excel(full_path, sheet_name='合并', index=False)
    state['save_path2'] = full_path
    var_save2.set(full_path)


# ══════════════════════════════════════════════
# 三、打开文件函数
# ══════════════════════════════════════════════

def open_book():
    if state['save_path']:
        os.startfile(state['save_path'])


def open_inducted():
    if state['save_path1']:
        os.startfile(state['save_path1'])


def open_merged():
    if state['save_path2']:
        os.startfile(state['save_path2'])


# ══════════════════════════════════════════════
# 四、核心数据处理函数（券商数据整理）
# ══════════════════════════════════════════════

def get_account_name(filepath: str) -> str:
    basename = os.path.basename(filepath)
    return os.path.splitext(basename)[0]


def process_broker_data(broker: str) -> pd.DataFrame:
    """读取并统一各券商对账单。

    2026-05-25 兼容增强：
    1. 广发常见字段可能叫“业务日期/业务标志名称/成交价格/清算金额”，也可能叫
       “成交日期/业务名称/成交均价/资金发生数/发生金额”；这里做别名自动识别。
    2. 不强制只读“对账单”sheet；优先读“对账单”，没有时依次尝试“交割单/资金流水/流水/明细”，最后读第一个 sheet。
    3. 证券代码统一清洗，避免 001335 变成 1335 或 1335.0。
    """
    src_path = state['path_src']
    fmt_path = state['path_fmt']

    name_map = pd.read_excel(fmt_path, usecols=[0, 1])
    date_col, biz_col, code_col, name_col, qty_col, price_col, amt_col, skip = BROKER_CONFIG[broker]

    # ── 1. 选择可用 sheet ──
    xls = pd.ExcelFile(src_path)
    preferred_sheets = ['对账单', '交割单', '资金流水', '流水', '明细', '交易流水']
    sheet_name = None
    for s in preferred_sheets:
        if s in xls.sheet_names:
            sheet_name = s
            break
    if sheet_name is None:
        sheet_name = xls.sheet_names[0]

    # 所有列都先按文本读取，后面再转数值，最大限度保留证券代码前导 0。
    book = pd.read_excel(src_path, sheet_name=sheet_name, skiprows=skip, dtype=str)
    book.columns = [str(c).strip() for c in book.columns]

    # 删除完全空行/空列
    book = book.dropna(how='all')
    book = book.loc[:, [c for c in book.columns if not str(c).startswith('Unnamed') or book[c].notna().any()]]

    for _, old_name, new_name in name_map.itertuples():
        book = book.replace(old_name, new_name)

    def _find_col(alias_list, required=True, logical_name=''):
        # 先精确匹配
        for a in alias_list:
            for c in book.columns:
                if str(c).strip() == a:
                    return c
        # 再包含匹配
        for a in alias_list:
            for c in book.columns:
                cs = str(c).strip()
                if a and (a in cs or cs in a):
                    return c
        if required:
            raise KeyError(f'未找到{logical_name or alias_list[0]}列，当前表头：{list(book.columns)}')
        return None

    # 各字段别名：兼容广发、华泰及常见券商格式
    date_alias  = [date_col, '业务日期', '成交日期', '交收日期', '发生日期', '清算日期', '日期', '交易日期']
    biz_alias   = [biz_col, '业务标志名称', '业务名称', '交易类别', '操作', '摘要', '业务摘要', '业务描述']
    code_alias  = [code_col, '证券代码', '证券编号', '股票代码', '证券代号']
    name_alias  = [name_col, '证券名称', '证券简称', '股票名称', '名称']
    qty_alias   = [qty_col, '成交数量', '成交股份', '发生数量', '股份发生数', '数量', '买卖数量']
    price_alias = [price_col, '成交均价', '成交价格', '成交价', '委托价格', '价格']
    amt_alias   = [amt_col, '发生金额', '清算金额', '资金发生数', '资金发生额', '成交金额', '金额']

    date_col2  = _find_col(date_alias,  True, '成交日期')
    biz_col2   = _find_col(biz_alias,   True, '业务名称')
    code_col2  = _find_col(code_alias,  True, '证券代码')
    name_col2  = _find_col(name_alias,  True, '证券名称')
    qty_col2   = _find_col(qty_alias,   True, '成交数量')
    price_col2 = _find_col(price_alias, True, '成交价格')
    amt_col2   = _find_col(amt_alias,   True, '发生金额/清算金额')

    base_cols = [date_col2, biz_col2, code_col2, name_col2, qty_col2, price_col2, amt_col2]
    extra_cols = [c for c in OPTIONAL_TXN_COLS if c in book.columns and c not in base_cols]
    # 广发有时“资金本次余额/资金余额”用于现金核对，能保留就保留
    for c in book.columns:
        cs = str(c).strip()
        if (('余额' in cs) or ('备注' in cs) or ('摘要' in cs) or ('市场' in cs)) and c not in base_cols and c not in extra_cols:
            extra_cols.append(c)

    book = book[base_cols + extra_cols].copy()
    rename_map = {
        date_col2: '成交日期', biz_col2: '业务名称', code_col2: '证券代码', name_col2: '证券名称',
        qty_col2: '成交数量', price_col2: '成交价格', amt_col2: '发生金额'
    }
    book = book.rename(columns=rename_map)
    book = book.loc[:, ~book.columns.duplicated()].copy()

    book.insert(0, column='账户名称', value=get_account_name(src_path))

    def _clean_code_for_book(v, market_value=''):
        """清洗证券代码，特别处理港股代码。

        关键规则：
        - 港股代码保持 5 位，如 6030 / 06030 -> 06030，700 -> 00700；
        - A股/ETF/北交所保持 6 位，如 6030 -> 006030；
        - 已带 HK/HK. 后缀/前缀的代码，按港股处理。
        """
        if v is None:
            return ''
        try:
            if pd.isna(v):
                return ''
        except Exception:
            pass

        raw = str(v).strip().replace('\t', '').replace(' ', '').upper()
        if not raw or raw in ('NAN', 'NONE'):
            return ''
        if raw.endswith('.0'):
            raw = raw[:-2]

        mkt = str(market_value or '').strip().upper()
        is_hk = (
            raw.startswith('HK') or raw.endswith('.HK') or
            '港' in mkt or mkt in ('HK', 'H股', '沪HK', '深HK') or 'HK' in mkt
        )

        # 去掉前后缀后只保留数字主体
        s = re.sub(r'\.[A-Z]+$', '', raw)
        s = re.sub(r'^(SH|SZ|BJ|NQ|HK)', '', s)

        if re.fullmatch(r'\d+', s):
            if is_hk:
                return s.zfill(5)[-5:]
            # 已经是 5 位且以 0 开头的，通常是港股原始代码，例如 06030，不能补成 006030。
            if len(s) == 5 and s.startswith('0'):
                return s
            if len(s) < 6:
                return s.zfill(6)
            return s

        return s

    def _to_num(x):
        if x is None:
            return 0.0
        s = str(x).strip().replace(',', '').replace('，', '').replace(' ', '')
        if s in ('', 'nan', 'None', '--'):
            return 0.0
        # 广发/华泰偶有括号负数
        neg = s.startswith('(') and s.endswith(')')
        s = s.strip('()')
        try:
            v = float(s)
            return -v if neg else v
        except Exception:
            return 0.0

    # 华泰等券商港股代码可能是 06030、00700；不能按 A 股逻辑补成 006030。
    # 如果原始表里有“交易市场/市场”等字段，则结合市场字段判断；没有市场字段时，5位且以0开头的代码按港股保留。
    market_col_for_code = None
    for _c in book.columns:
        _cs = str(_c).strip()
        if _c != '交易市场' and (('市场' in _cs) or _cs in ('交易市场', '市场类别')):
            market_col_for_code = _c
            break
    if '交易市场' in book.columns:
        market_col_for_code = '交易市场'

    if market_col_for_code:
        book['证券代码'] = book.apply(lambda r: _clean_code_for_book(r.get('证券代码'), r.get(market_col_for_code)), axis=1)
    else:
        book['证券代码'] = book['证券代码'].apply(_clean_code_for_book)
    book['成交数量'] = book['成交数量'].apply(_to_num).abs()
    book['成交价格'] = book['成交价格'].apply(_to_num)
    # 交易整理表中的发生金额仍取绝对值；现金/资金投入方向由业务名称规则判断。
    book['发生金额'] = book['发生金额'].apply(_to_num).abs()

    return book

def process_and_preview():
    text_pad.config(state=tk.NORMAL)
    broker = var_broker.get()

    if not broker:
        messagebox.showerror('券商选择错误', '请点击下拉框选择券商')
        return

    if broker not in state['path_src']:
        messagebox.showerror('文件选择错误', '请确认所选源文件与所选券商是否对应')
        return

    try:
        state['book'] = process_broker_data(broker)
        text_pad.delete(1.0, tk.END)
        text_pad.insert('insert', state['book'].to_string())
        update_status(f'✅ 数据整理完成，共 {len(state["book"])} 条记录')
    except KeyError:
        messagebox.showerror('出错了', '请检查导入的券商数据表格表头名称是否是标准格式')
    except Exception as e:
        messagebox.showerror('出错了', f'处理失败，错误信息：{e}')


def merge_files():
    text_pad.config(state=tk.NORMAL)
    merge_dir = state['path_merge']
    if not merge_dir:
        messagebox.showerror('错误', '请先选择待合并文件夹')
        return
    try:
        frames = []
        for root_dir, _, files in os.walk(merge_dir):
            for file in files:
                file_path = os.path.join(root_dir, file)
                df = pd.read_excel(file_path, engine='openpyxl', converters={'证券代码': str})
                frames.append(df)

        if not frames:
            messagebox.showwarning('提示', '所选文件夹内没有找到文件')
            return

        state['df_merged'] = pd.concat(frames, axis=0, ignore_index=True)
        text_pad.delete(1.0, tk.END)
        text_pad.insert('insert', state['df_merged'].to_string())
        update_status(f'✅ 文件合并完成，共合并 {len(frames)} 个文件，{len(state["df_merged"])} 条记录')
    except Exception as e:
        messagebox.showerror('出错了',
            f'合并失败：{e}\n请确认文件夹内的文件都已经过程序处理，且仅包含待合并文件。')


def induct_data():
    text_pad.config(state=tk.NORMAL)
    induct_path = state['path_induct']
    if not induct_path:
        messagebox.showerror('错误', '请先选择待归纳文件')
        return

    try:
        data = pd.read_excel(induct_path,
                             converters={'证券代码': str, '证券名称': str},
                             index_col=0)

        data_valid = data[data['发生金额'] != 0]
        data_valid = data_valid.dropna(subset=['证券代码'])
        data_valid = data_valid[~data_valid['证券名称'].isin(['GC001'])]

        code_list = data_valid['证券代码'].unique()
        rows = []
        for code in code_list:
            sub = data_valid[data_valid['证券代码'] == code]
            buy  = sub[sub['业务名称'] == '证券买入']
            buy_amt = buy['发生金额'].sum()
            buy_qty = buy['成交数量'].sum()
            sell = sub[sub['业务名称'] == '证券卖出']
            sell_amt = sell['发生金额'].sum()
            sell_qty = sell['成交数量'].sum()
            div  = sub[sub['业务名称'] == '分红']['发生金额'].sum()
            lend = sub[sub['业务名称'] == '转融券出借利息']['发生金额'].sum()
            rows.append([code, buy_qty, buy_amt, sell_qty, sell_amt, div, lend])

        gn = pd.DataFrame(rows,
                          columns=['证券代码', '买入数量', '买入金额合计',
                                   '卖出数量', '卖出金额合计', '分红', '转融券出借利息'])

        gn.insert(3, '买入均价', (gn['买入金额合计'] / gn['买入数量']).round(2))
        gn.insert(6, '卖出均价', (gn['卖出金额合计'] / gn['卖出数量']).round(2))

        code_to_name = (data_valid[['证券代码', '证券名称']]
                        .drop_duplicates(subset='证券代码')
                        .set_index('证券代码')['证券名称'])
        gn.insert(1, '证券名称', gn['证券代码'].map(code_to_name))

        data_no_code = data[data['证券代码'].isna()]
        unnamed_lend = round(data_no_code[data_no_code['业务名称'] == '转融券出借利息']['发生金额'].sum(), 2)
        unnamed_div  = data_no_code[data_no_code['业务名称'] == '分红']['发生金额'].sum()

        extra = pd.DataFrame({
            '证券名称':     ['未记名转融券出借利息合计', '未记名分红合计'],
            '分红':         [None, unnamed_div],
            '转融券出借利息': [unnamed_lend, None],
        })
        gn = pd.concat([gn, extra], join='outer', axis=0, ignore_index=True)

        state['df_inducted'] = gn
        text_pad.delete(1.0, tk.END)
        text_pad.insert('insert', gn.to_string())
        update_status(f'✅ 归纳汇总完成，共 {len(code_list)} 只证券')

    except Exception as e:
        messagebox.showerror('出错了',
            f'归纳失败：{e}\n请确认待归纳文件已经过整理或合并程序处理。')


# ══════════════════════════════════════════════
# 五、净值计算函数（整合自 calc_nav.py）
# ══════════════════════════════════════════════

def load_business_rules(filepath):
    """读取业务变化逻辑规则 → {业务名称: (数量变化, 金额变化)}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']
    rules = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[1]).strip() if row[1] else ''
        if name:
            rules[name] = (
                str(row[2]).strip() if row[2] else '不变',
                str(row[5]).strip() if row[5] else '不变',
            )
    wb.close()
    return rules


def load_initial_positions(filepath):
    """
    读取期初持仓。
    返回: positions {代码: (名称, 数量, 成本价)}, cash, hkd_rates, initial_total
    """
    wb      = openpyxl.load_workbook(filepath, data_only=True)
    wb_fml  = openpyxl.load_workbook(filepath)
    ws      = wb['期初']
    ws_f    = wb_fml['期初']
    positions     = {}
    cash          = 0.0
    holding_value = 0.0

    for r in range(2, ws.max_row + 1):
        row   = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        row_f = next(ws_f.iter_rows(min_row=r, max_row=r, values_only=True))
        code, name, qty = row[0], row[1], row[2]
        col_e, col_f, col_e_f = row[4], row[5], row_f[4]

        if code is None and name == '现金':
            cash = float(col_e) if col_e else 0.0
            if cash == 0.0 and col_e_f is not None:
                nums = re.findall(r'[\d.]+', str(col_e_f))
                if nums:
                    cash = sum(float(n) for n in nums)
        elif code is not None and qty is not None:
            code_s = str(code).strip()
            cost   = float(col_f) if col_f is not None else 0.0
            positions[code_s] = (str(name).strip() if name else code_s, int(qty), cost)
            holding_value += int(qty) * cost

    # 港股通汇率
    ws2 = wb['港股通汇率']
    hkd_rates = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        dt, rate = row[0], row[3]
        if dt and rate:
            if isinstance(dt, (datetime, date)):
                key = dt.strftime('%Y-%m-%d')
            else:
                key = str(dt).strip()
            hkd_rates[key] = float(rate)

    wb.close(); wb_fml.close()
    return positions, cash, hkd_rates, holding_value + cash


def load_transactions_from_df(df):
    """
    从 DataFrame（一键处理生成的 state['book']）读取交易流水 → 按日期排序的列表。
    列顺序：账户名称, 成交日期, 业务名称, 证券代码, 证券名称, 成交数量, 成交价格, 发生金额
    """
    txns = []
    for _, row in df.iterrows():
        raw_date = row['成交日期']

        # 兼容华泰/广发等对账单：process_broker_data 为保留证券代码前导 0，
        # 会用 dtype=str 读取整张表，因此成交日期常变成字符串，如：
        #   20250102 / 2025-01-02 / 2025/1/2 / 2025-01-02 00:00:00
        # 旧逻辑只识别 datetime、date、int、float，遇到字符串日期会全部跳过，
        # 导致 transactions 为空并弹出“交易流水为空”。
        def _parse_trade_date(v):
            if v is None:
                return None
            try:
                if pd.isna(v):
                    return None
            except Exception:
                pass
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            if isinstance(v, (int, float)):
                try:
                    return datetime.strptime(str(int(v)), '%Y%m%d').date()
                except Exception:
                    return None
            s = str(v).strip()
            if not s or s.lower() in ('nan', 'nat', 'none'):
                return None
            # Excel 文本日期常见形式：20250102、20250102.0、2025-01-02、2025/1/2
            s0 = s.replace('年', '-').replace('月', '-').replace('日', '')
            s0 = s0.split()[0].strip()
            if s0.endswith('.0'):
                s0 = s0[:-2]
            digits = re.sub(r'\D', '', s0)
            if len(digits) == 8:
                try:
                    return datetime.strptime(digits, '%Y%m%d').date()
                except Exception:
                    pass
            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
                try:
                    return datetime.strptime(s0, fmt).date()
                except Exception:
                    pass
            try:
                return pd.to_datetime(s, errors='coerce').date()
            except Exception:
                return None

        d = _parse_trade_date(raw_date)
        if d is None:
            continue

        code_s = normalize_sec_code(row['证券代码']) if row['证券代码'] is not None else ''
        # 8开头6位 → 北交所代码转换
        if len(code_s) == 6 and code_s.startswith('8'):
            code_s = '920' + code_s[3:]

        def _clean_extra(col_name):
            if col_name not in row.index:
                return ''
            v = row[col_name]
            if v is None or (pd.notna(v) is False):
                return ''
            sv = str(v).strip()
            return '' if sv.lower() in ('nan', 'nat', 'none') else sv

        txns.append({
            'date':     d,
            'biz_name': str(row['业务名称']).strip() if row['业务名称'] is not None and str(row['业务名称']).strip() != 'nan' else '',
            'sec_code': code_s,
            'sec_name': str(row['证券名称']).strip() if row['证券名称'] is not None and str(row['证券名称']).strip() != 'nan' else '',
            'qty':      int(row['成交数量'])   if row['成交数量']  is not None and pd.notna(row['成交数量'])  else 0,
            'price':    float(row['成交价格']) if row['成交价格']  is not None and pd.notna(row['成交价格'])  else 0.0,
            'amount':   float(row['发生金额']) if row['发生金额']  is not None and pd.notna(row['发生金额'])  else 0.0,
            'remark':   _clean_extra('备注'),
            'summary':  _clean_extra('摘要'),
            'operation': _clean_extra('操作'),
            'market':   _clean_extra('交易市场') or _clean_extra('市场'),
        })
    txns.sort(key=lambda x: x['date'])
    return txns


def build_ipo_code_set(transactions):
    """收集新股相关证券代码，用于识别后续“新股托管转入/限售转无限售”。"""
    ipo_codes = set()
    for txn in transactions:
        biz = str(txn.get('biz_name', '')).strip()
        code = normalize_sec_code(txn.get('sec_code', ''))
        text = ''.join(str(txn.get(k, '') or '') for k in ('biz_name', 'remark', 'summary', 'operation'))
        if code and ('新股' in text or 'XGJX' in text.upper()):
            ipo_codes.add(code)
        if code and biz in ('新股入账', '新股入帐', '新股申购', '申购中签'):
            ipo_codes.add(code)
    return ipo_codes


def is_new_stock_custody_transfer(txn, ipo_codes):
    """
    只有新股相关的“托管转入”才影响持仓。

    万联流水里新股限售转无限售常见备注为“限售股份转无限售-XGJX”。
    普通托管转入/转托管不应改变本账户真实持仓，所以不能一律按数量增加。
    """
    code = normalize_sec_code(txn.get('sec_code', ''))
    text = ''.join(str(txn.get(k, '') or '') for k in ('biz_name', 'sec_name', 'remark', 'summary', 'operation'))
    text_upper = text.upper()
    if 'XGJX' in text_upper:
        return True
    if '新股' in text:
        return True
    if code and code in ipo_codes:
        return True
    return False


def _convertible_root(name):
    """提取配债/转债的同一主体名称，如“鼎龙配债”“鼎龙转债”都返回“鼎龙”。"""
    s = str(name or '').strip()
    if not s or s.lower() in ('nan', 'none'):
        return ''
    s = re.sub(r'(配债|转债|发债|可转债)$', '', s)
    return s.strip()


def build_convertible_bond_alias_map(transactions):
    """把配债权利代码映射到正式转债代码。

    例：380054 鼎龙配债 -> 123255 鼎龙转债。配债缴款发生时应形成
    正式转债持仓成本，而不是留在配债权利代码上。
    """
    formal_by_root = {}
    allot_by_root = {}
    for txn in transactions or []:
        code = normalize_sec_code(txn.get('sec_code', ''))
        name = str(txn.get('sec_name', '') or '').strip()
        root = _convertible_root(name)
        if not root or not code:
            continue
        if '转债' in name:
            formal_by_root[root] = code
        elif '配债' in name:
            allot_by_root[root] = code

    alias = {}
    for root, allot_code in allot_by_root.items():
        formal_code = formal_by_root.get(root)
        if formal_code:
            alias[allot_code] = formal_code
    return alias


def is_convertible_allotment_payment(txn):
    """识别可转债配债缴款流水。"""
    biz = str(txn.get('biz_name', '') or '')
    name = str(txn.get('sec_name', '') or '')
    text = biz + name + str(txn.get('remark', '') or '') + str(txn.get('summary', '') or '')
    return '配债' in text and ('缴款' in text or '买入' in biz)


def is_convertible_allotment_intermediate(txn):
    """识别配债过程中的权利上账/上市修正流水，避免重复形成持仓。"""
    biz = str(txn.get('biz_name', '') or '')
    text = (
        biz
        + str(txn.get('sec_name', '') or '')
        + str(txn.get('remark', '') or '')
        + str(txn.get('summary', '') or '')
        + str(txn.get('operation', '') or '')
    )
    if '配债' not in text:
        return False
    return any(k in biz or k in text for k in ('权证上账', '股份上市', '交收股份修正'))


def load_bank_transfers(transactions):
    """
    从流水中提取银行转存/转取记录，用于资金投入计算。

    统一口径：
      - 银行转存 = 资金投入，金额取正数
      - 银行转取 = 资金收回，金额取负数

    注意：不同券商流水里“发生金额”可能本身带正负号，
    因此这里按业务名称强制规范符号，避免“银行转取”负数再取反导致变成正投入。
    """
    result = []
    for txn in transactions:
        biz = str(txn.get('biz_name', '')).strip()
        amount_raw = float(txn.get('amount') or 0.0)
        if biz == '银行转存':
            result.append({'date': txn['date'], 'amount': abs(amount_raw), 'biz_name': biz})
        elif biz == '银行转取':
            result.append({'date': txn['date'], 'amount': -abs(amount_raw), 'biz_name': biz})
    result.sort(key=lambda x: x['date'])
    return result


def build_capital_events_from_bank_transfers(initial_total, bank_transfers, start_date):
    """
    根据期初总资产 + 流水中的银行转存/转取生成资金投入事件。

    使用与“资金投入/资金占用”表一致的投入积数逻辑：
      天数 = 估值日 - 操作日期 + 1
      投入积数 = 资金净投入 × 天数
      平均投入/资金占用 = 投入积数合计 ÷ 估值期间总天数

    不再读取《证券账户持仓表》的“资金投入”sheet。
    """
    if isinstance(start_date, datetime):
        start_date = start_date.date()

    events = []
    init_amt = float(initial_total or 0.0)
    if abs(init_amt) > 1e-9:
        events.append({
            'date': start_date,
            'amount': init_amt,
            'biz_name': '期初总资产',
            'source': '期初总资产'
        })

    for tf in bank_transfers or []:
        d = tf.get('date')
        if isinstance(d, datetime):
            d = d.date()
        if d is None or d < start_date:
            continue
        amount = float(tf.get('amount') or 0.0)
        if abs(amount) < 1e-9:
            continue
        events.append({
            'date': d,
            'amount': amount,
            'biz_name': tf.get('biz_name', '银行转账'),
            'source': '交易流水银行转存/转取'
        })

    events.sort(key=lambda x: x['date'])
    return events



def _safe_float(v):
    """把金额/数量类值安全转 float。"""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    s = str(v).strip().replace(',', '').replace('，', '')
    if not s or s.lower() in ('nan', 'none', 'nat') or s in ('-', '—', '--'):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _has_sheet(filepath, sheet_name):
    """判断工作簿是否包含指定 sheet。"""
    if not filepath or not os.path.exists(filepath):
        return False
    try:
        xls = pd.ExcelFile(filepath)
        return sheet_name in xls.sheet_names
    except Exception:
        return False


def resolve_capital_input_file(base_file=None):
    """查找带“资金投入”sheet 的证券账户持仓表。"""
    if base_file and _has_sheet(base_file, '资金投入'):
        return base_file

    dirs = []
    if base_file:
        dirs.append(os.path.dirname(os.path.abspath(base_file)))
    try:
        dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass
    dirs.append(os.getcwd())

    seen = set()
    for d in dirs:
        if not d or d in seen or not os.path.isdir(d):
            continue
        seen.add(d)
        try:
            names = os.listdir(d)
        except Exception:
            continue
        candidates = []
        for name in names:
            low = name.lower()
            if ('证券账户持仓表' in name or '持仓表' in name) and low.endswith(('.xls', '.xlsx', '.xlsm')):
                candidates.append(os.path.join(d, name))
        for path in sorted(candidates, reverse=True):
            if _has_sheet(path, '资金投入'):
                return path
    return None


def _make_account_hint(init_path):
    """从期初文件名中提取账户名提示，例如 万联正涵250101_251231 -> 万联正涵。"""
    base = os.path.splitext(os.path.basename(str(init_path or '')))[0]
    base = re.sub(r'\d{6,8}_\d{6,8}$', '', base)
    base = re.sub(r'\d{6,8}$', '', base)
    return base.strip()


def _norm_name_for_match(s):
    return re.sub(r'\s+', '', str(s or '').replace('\n', '').replace('（新）', '').strip())


def load_capital_events_from_fund_sheet(filepath, account_hint=''):
    """
    读取《证券账户持仓表》里的“资金投入”sheet，并转换成投入事件。

    资金占用逻辑来自表格：
        天数 = 估值日 - 操作日期 + 1
        投入积数 = 资金净投入 × 天数
        资金占用/平均投入 = 投入积数合计 ÷ 估值期间总天数

    如果能在账户明细列中匹配到当前账户（如“万联正涵”），就使用该账户列；
    匹配不到时，退回使用左侧“资金净投入”汇总列。
    """
    if not filepath or not os.path.exists(filepath):
        return [], '', filepath

    try:
        raw = pd.read_excel(filepath, sheet_name='资金投入', header=None, dtype=object)
    except Exception:
        return [], '', filepath

    if raw.empty:
        return [], '', filepath

    hint = _norm_name_for_match(account_hint)

    title_row = None
    for i in range(min(len(raw), 30)):
        v = str(raw.iat[i, 0]).strip() if pd.notna(raw.iat[i, 0]) else ''
        if '证券户资金投入' in v:
            title_row = i
            break
    if title_row is None:
        title_row = 0

    header_row = None
    for i in range(title_row, min(len(raw), title_row + 15)):
        vals = [_norm_name_for_match(x) for x in raw.iloc[i].tolist()]
        if '操作日期' in vals and '资金净投入' in vals:
            header_row = i
            break
    if header_row is None:
        return [], '', filepath

    account_row = header_row + 1
    account_col = None
    account_name = ''
    if account_row < len(raw) and hint:
        for c in range(8, raw.shape[1]):
            nm = _norm_name_for_match(raw.iat[account_row, c])
            if nm and (nm == hint or nm in hint or hint in nm):
                account_col = c
                account_name = str(raw.iat[account_row, c]).replace('\n', '').strip()
                break

    if account_col is None:
        try:
            broker_hint = _norm_name_for_match(var_broker.get())
        except Exception:
            broker_hint = ''
        if account_row < len(raw) and broker_hint:
            for c in range(8, raw.shape[1]):
                nm = _norm_name_for_match(raw.iat[account_row, c])
                if nm and broker_hint in nm:
                    account_col = c
                    account_name = str(raw.iat[account_row, c]).replace('\n', '').strip()
                    break

    if account_col is None:
        account_col = 3
        account_name = '资金净投入汇总列'

    events = []
    for r in range(header_row + 2, len(raw)):
        first = raw.iat[r, 0] if raw.shape[1] > 0 else None
        first_s = str(first).strip() if first is not None and not pd.isna(first) else ''
        if first_s in ('期间投入', '合计', '总资产') or '正涵1号' in first_s:
            break
        d = parse_excel_date(first)
        if d is None:
            continue
        amount = _safe_float(raw.iat[r, account_col])
        if amount is None or abs(amount) < 1e-9:
            continue
        events.append({'date': d, 'amount': float(amount), 'row': r + 1, 'source': account_name})

    events.sort(key=lambda x: x['date'])
    return events, account_name, filepath


def compute_capital_from_events(events, target_date, start_date):
    """
    按资金投入表的资金占用逻辑计算：
        period_days = target_date - start_date + 1
        days_left   = target_date - event_date + 1
        cap_sum     = Σ event_amount × days_left
        avg_cap     = cap_sum / period_days
        net_capital = Σ event_amount
    """
    if isinstance(target_date, datetime):
        target_date = target_date.date()
    if isinstance(start_date, datetime):
        start_date = start_date.date()
    period_days = max((target_date - start_date).days + 1, 1)
    cap_sum = 0.0
    net_capital = 0.0
    for ev in events or []:
        d = ev.get('date')
        amount = float(ev.get('amount') or 0.0)
        if d is None or d > target_date:
            continue
        days_left = (target_date - d).days + 1
        if days_left > 0:
            cap_sum += amount * days_left
            net_capital += amount
    avg_cap = cap_sum / period_days if period_days > 0 else cap_sum
    return cap_sum, period_days, avg_cap, net_capital


def normalize_sec_code(code):
    """统一证券代码格式：兼容 002714.SZ / 600000.SH / 430607.BJ / 430607.NQ / HK02269 / 02269。"""
    if code is None:
        return ''
    s = str(code).strip()
    if not s or s.lower() == 'nan':
        return ''
    if s.endswith('.0'):
        s = s[:-2]
    s = s.upper().strip()

    # 去掉市场后缀：002714.SZ、600000.SH、430607.BJ、430607.NQ、02269.HK 等
    # 关键修复：股票池2025年收盘价.xlsx 里大树智能是 430607.NQ，原代码只去掉 SH/SZ/BJ/HK，
    # 会把代码保存成 430607.NQ，估值时用 430607 查不到。这里统一去掉点号后的全部市场后缀。
    s = re.sub(r'\.[A-Z]+$', '', s)
    # 去掉市场前缀：SH600000、SZ000001、BJ430607、NQ430607、HK02269
    s = re.sub(r'^(SH|SZ|BJ|NQ|HK)', '', s)
    s = s.strip()

    # 保留纯数字。港股通常 5 位，A/B/北交所通常 6 位。
    if s.isdigit():
        if len(s) < 5:
            s = s.zfill(5)
        elif len(s) == 5:
            s = s.zfill(5)
        elif len(s) == 6:
            s = s.zfill(6)
    return s


def is_reverse_repo_code(code):
    """逆回购品种按面值 100 估值。"""
    code_s = normalize_sec_code(code)
    return bool(code_s) and (code_s.startswith('204') or code_s == '131810')


def parse_excel_date(v):
    """把 Excel 日期/字符串/20250103 数字统一转 date。"""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'):
        return None
    try:
        # 20250103 / 20250103.0
        if re.fullmatch(r'\d{8}(\.0)?', s):
            return datetime.strptime(s[:8], '%Y%m%d').date()
        dt = pd.to_datetime(s, errors='coerce')
        if pd.isna(dt):
            return None
        return dt.date()
    except Exception:
        return None

def resolve_existing_file(preferred_path, base_file=None, candidates=None):
    """
    兼容原来的硬编码路径：
    1）优先使用 preferred_path；
    2）再到期初文件同目录找候选文件名；
    3）最后到程序所在目录找候选文件名。
    """
    if preferred_path and os.path.exists(preferred_path):
        return preferred_path

    candidates = candidates or []
    dirs = []
    if base_file:
        dirs.append(os.path.dirname(os.path.abspath(base_file)))
    try:
        dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass
    dirs.append(os.getcwd())

    seen = set()
    for d in dirs:
        if not d or d in seen:
            continue
        seen.add(d)
        for name in candidates:
            path = os.path.join(d, name)
            if os.path.exists(path):
                return path
    return preferred_path



def load_stock_pool_close_prices(filepath):
    """
    读取本地收盘价表，兼容“通达信提取收盘价结果.xlsx”和“股票池2025年收盘价.xlsx”。

    支持格式：
        第1行可能是“收盘价(元)”说明行；
        真正表头行为：代码 | 简称 | 2025-01-01 | 2025-01-02 | ...
    返回：{('YYYY-MM-DD', '证券代码'): 收盘价}
    """
    if not filepath or not os.path.exists(filepath):
        return {}

    prices = {}

    def _to_float_price(v):
        if v is None or pd.isna(v):
            return None
        s = str(v).strip().replace(',', '')
        if not s or s.lower() == 'nan' or s in ('--', '-', '—', '无', 'None', 'NONE'):
            return None
        try:
            x = float(s)
        except Exception:
            return None
        return x if x > 0 else None

    def _save_price(d, code, close):
        d = parse_excel_date(d)
        code = normalize_sec_code(code)
        close = _to_float_price(close)
        if d is None or not code or close is None:
            return
        prices[(d.strftime('%Y-%m-%d'), code)] = close

    try:
        xls = pd.ExcelFile(filepath)
    except Exception:
        return prices

    for sheet_name in xls.sheet_names:
        try:
            raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None, dtype=object)
            raw = raw.dropna(how='all').dropna(axis=1, how='all')
        except Exception:
            continue
        if raw.empty:
            continue

        # 格式A：股票池矩阵。某一行含“代码/证券代码”，右侧为日期列。
        header_row = None
        code_col_idx = None
        date_cols = []
        scan_rows = min(15, len(raw))
        for r in range(scan_rows):
            row_vals = [str(x).strip() if not pd.isna(x) else '' for x in raw.iloc[r].tolist()]
            for c, val in enumerate(row_vals):
                if val in ('代码', '证券代码'):
                    tmp_date_cols = []
                    for j in range(c + 1, raw.shape[1]):
                        if parse_excel_date(raw.iat[r, j]) is not None:
                            tmp_date_cols.append(j)
                    if tmp_date_cols:
                        header_row = r
                        code_col_idx = c
                        date_cols = tmp_date_cols
                        break
            if header_row is not None:
                break

        if header_row is not None and code_col_idx is not None and date_cols:
            for i in range(header_row + 1, raw.shape[0]):
                code = raw.iat[i, code_col_idx]
                code_s = normalize_sec_code(code)
                if not code_s:
                    continue
                for j in date_cols:
                    _save_price(raw.iat[header_row, j], code_s, raw.iat[i, j])
            continue

        # 格式B：普通明细表。日期 | 证券代码/代码 | 收盘价
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            df = df.dropna(how='all').dropna(axis=1, how='all')
        except Exception:
            continue
        if df.empty:
            continue
        cols_str = [str(c).strip() for c in df.columns]
        date_col = code_col = close_col = None
        for c, cs in zip(df.columns, cols_str):
            low = cs.lower()
            if date_col is None and (cs in ('日期', '交易日期') or '日期' in cs or low == 'date'):
                date_col = c
            if code_col is None and (cs in ('证券代码', '代码') or '证券代码' in cs or low == 'code'):
                code_col = c
            if close_col is None and (cs in ('收盘价', '收盘') or '收盘' in cs or low in ('close', 'close_price')):
                close_col = c
        if date_col is not None and code_col is not None and close_col is not None:
            for _, row in df.iterrows():
                _save_price(row.get(date_col), row.get(code_col), row.get(close_col))
            continue

        # 格式C：日期横向表。日期 | 000001 | 000651 | 430607 | 02269 ...
        date_col = df.columns[0]
        for c, cs in zip(df.columns, cols_str):
            if cs in ('日期', '交易日期') or '日期' in cs or cs.lower() == 'date':
                date_col = c
                break
        for _, row in df.iterrows():
            d = parse_excel_date(row.get(date_col))
            if d is None:
                continue
            for col in df.columns:
                if col == date_col:
                    continue
                _save_price(d, col, row.get(col))

    return prices


def find_local_close_price(local_prices, code, target_date, max_lookback=240):
    """从本地价格表中查找目标日或之前最近可用收盘价/净值。"""
    code_s = normalize_sec_code(code)
    if not code_s:
        return None
    if is_reverse_repo_code(code_s):
        return 100.0
    if isinstance(target_date, datetime):
        target_date = target_date.date()

    candidates = []
    def add(x):
        x = normalize_sec_code(x)
        if x and x not in candidates:
            candidates.append(x)

    add(code_s)
    if code_s.isdigit():
        if len(code_s) == 5:
            add('HK' + code_s)
        if len(code_s) == 6:
            add(code_s.zfill(6))
            if code_s.startswith('920'):
                add('430' + code_s[-3:])
            if code_s.startswith('430'):
                add('920' + code_s[-3:])
            if code_s.startswith('8'):
                add('920' + code_s[3:])

    for delta in range(int(max_lookback) + 1):
        d_s = (target_date - timedelta(days=delta)).strftime('%Y-%m-%d')
        for c in candidates:
            v = local_prices.get((d_s, c))
            if v is not None:
                return v
    return None

class LocalPriorityClosePriceProvider:
    """
    本地价格提供器：通达信收盘价优先，股票池收盘价/私募净值兜底。

    取价顺序：
    1）先查“通达信提取收盘价结果.xlsx”；
    2）找不到再查“股票池2025年收盘价.xlsx”；
    3）两边都没有，写入“缺失价格”sheet。
    """
    def __init__(self, primary_prices=None, fallback_prices=None, primary_file=None, fallback_file=None, lookback_days=240, log_func=None):
        self.primary_prices = primary_prices or {}
        self.fallback_prices = fallback_prices or {}
        self.primary_file = primary_file
        self.fallback_file = fallback_file
        self.lookback_days = int(lookback_days)
        self.log_func = log_func
        self.missing = []
        self.primary_hit_count = 0
        self.fallback_hit_count = 0

    def save_cache(self):
        # 兼容主流程里的 save_cache 调用；本地价格无需保存缓存。
        return None

    def get_close_price(self, code, target_date, name='', max_lookback=None):
        lookback = int(max_lookback or self.lookback_days)
        d = target_date.date() if isinstance(target_date, datetime) else target_date

        # 204001/GC001、131810/R-001 这类逆回购按 100 估值。
        if is_reverse_repo_code(code):
            return 100.0

        v = find_local_close_price(self.primary_prices, code, d, max_lookback=lookback)
        if v is not None:
            self.primary_hit_count += 1
            return v

        v = find_local_close_price(self.fallback_prices, code, d, max_lookback=lookback)
        if v is not None:
            self.fallback_hit_count += 1
            return v

        # 不在这里立即写缺失。后续还有“新股入账价/发行价”等兜底逻辑。
        # 真正缺失会在 calc_position_value() 确认所有兜底都失败后再记录，
        # 避免明细已估值、但“缺失价格”sheet 仍显示缺失。
        return None


class OnlineClosePriceProvider:
    """
    联网收盘价提供器。

    数据源：AKShare 东方财富历史行情接口。
    - A股/ETF/北交所：ak.stock_zh_a_hist
    - 港股：ak.stock_hk_hist

    特点：
    1）不再依赖本地收盘价 Excel；
    2）目标日没有行情时，自动向前找最近一个交易日；
    3）带 JSON 本地缓存，避免重复联网请求；
    4）失败时记录缺失明细，不中断整批净值计算。
    """
    def __init__(self, cache_file=None, lookback_days=240, retry=3, sleep_sec=0.35, log_func=None):
        self.cache_file = cache_file or os.path.join(os.getcwd(), 'online_price_cache.json')
        self.lookback_days = int(lookback_days)
        self.retry = int(retry)
        self.sleep_sec = float(sleep_sec)
        self.log_func = log_func
        self.cache = {}
        self.missing = []
        self._ak = None
        self._load_cache()

    def _log(self, msg):
        if self.log_func:
            try:
                self.log_func(msg)
            except Exception:
                pass

    def _load_cache(self):
        try:
            if self.cache_file and os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    self.cache = data
        except Exception:
            self.cache = {}

    def save_cache(self):
        try:
            if not self.cache_file:
                return
            os.makedirs(os.path.dirname(os.path.abspath(self.cache_file)), exist_ok=True)
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_akshare(self):
        if self._ak is not None:
            return self._ak
        try:
            import akshare as ak
        except Exception as e:
            raise RuntimeError(
                '未安装 AKShare，无法联网提取收盘价。请先在命令行执行：pip install akshare -U'
            ) from e
        self._ak = ak
        return ak

    def _cache_key(self, market, code, d):
        if isinstance(d, (datetime, date)):
            d = d.strftime('%Y-%m-%d')
        return f'{market}:{normalize_sec_code(code)}:{d}'

    def _infer_market(self, code, name=''):
        code_s = normalize_sec_code(code)
        name_s = str(name or '')
        if not code_s:
            return 'A'
        # 港股通常是 5 位数字；本程序的 A 股/ETF/北交所均按 6 位处理。
        if len(code_s) == 5 and code_s.isdigit():
            return 'HK'
        if '港' in name_s or name_s.upper().startswith('HK'):
            return 'HK'
        return 'A'

    def _candidate_codes(self, code, market):
        code_s = normalize_sec_code(code)
        out = []
        def add(x):
            x = normalize_sec_code(x)
            if x and x not in out:
                out.append(x)
        add(code_s)
        if market == 'A' and code_s.isdigit() and len(code_s) == 6:
            # 北交所/老三板口径兼容：430607 与 920607 互相兜底；8开头旧代码转920口径。
            if code_s.startswith('430'):
                add('920' + code_s[-3:])
            if code_s.startswith('920'):
                add('430' + code_s[-3:])
            if code_s.startswith('8'):
                add('920' + code_s[3:])
        return out

    def _to_float_close(self, v):
        if v is None or pd.isna(v):
            return None
        s = str(v).strip().replace(',', '')
        if not s or s.lower() == 'nan' or s in ('--', '-', '—', 'None', '无'):
            return None
        try:
            x = float(s)
        except Exception:
            return None
        return x if x > 0 else None

    def _fetch_history(self, market, code, start_date, end_date):
        ak = self._load_akshare()
        start_s = start_date.strftime('%Y%m%d')
        end_s = end_date.strftime('%Y%m%d')

        last_err = None
        for attempt in range(1, self.retry + 1):
            try:
                if market == 'HK':
                    df = ak.stock_hk_hist(symbol=normalize_sec_code(code), period='daily',
                                          start_date=start_s, end_date=end_s, adjust='')
                else:
                    df = ak.stock_zh_a_hist(symbol=normalize_sec_code(code), period='daily',
                                            start_date=start_s, end_date=end_s, adjust='')
                if df is None or df.empty:
                    return pd.DataFrame()
                return df
            except Exception as e:
                last_err = e
                time.sleep(self.sleep_sec * attempt)
        raise last_err

    def _put_df_to_cache(self, market, code, df):
        if df is None or df.empty:
            return
        date_col = None
        close_col = None
        for c in df.columns:
            cs = str(c).strip()
            low = cs.lower()
            if date_col is None and (cs in ('日期', '交易日期') or low in ('date', 'trade_date')):
                date_col = c
            if close_col is None and (cs in ('收盘', '收盘价') or low in ('close', 'close_price')):
                close_col = c
        if date_col is None or close_col is None:
            return
        code_s = normalize_sec_code(code)
        for _, row in df.iterrows():
            d = parse_excel_date(row.get(date_col))
            close = self._to_float_close(row.get(close_col))
            if d and close is not None:
                self.cache[self._cache_key(market, code_s, d)] = close

    def get_close_price(self, code, target_date, name='', max_lookback=None):
        """返回目标日或目标日前最近交易日收盘价。"""
        code_s = normalize_sec_code(code)
        if not code_s:
            return None
        if is_reverse_repo_code(code_s):
            return 100.0

        if isinstance(target_date, datetime):
            target_date = target_date.date()
        lookback = int(max_lookback or self.lookback_days)
        market = self._infer_market(code_s, name)

        # 先从缓存向前找，命中即返回。
        for c in self._candidate_codes(code_s, market):
            for delta in range(lookback + 1):
                d = target_date - timedelta(days=delta)
                key = self._cache_key(market, c, d)
                if key in self.cache:
                    return self.cache[key]

        # 缓存没有，再联网拉一段历史数据。
        start_date = target_date - timedelta(days=lookback)
        end_date = target_date + timedelta(days=1)
        errors = []
        for c in self._candidate_codes(code_s, market):
            try:
                df = self._fetch_history(market, c, start_date, end_date)
                self._put_df_to_cache(market, c, df)
                self.save_cache()
            except Exception as e:
                errors.append(f'{c}: {e}')

            for delta in range(lookback + 1):
                d = target_date - timedelta(days=delta)
                key = self._cache_key(market, c, d)
                if key in self.cache:
                    return self.cache[key]

        self.missing.append({
            'date': target_date.strftime('%Y-%m-%d'),
            'code': code_s,
            'name': str(name or ''),
            'market': market,
            'error': '; '.join(errors[-2:]) if errors else '接口返回空数据',
        })
        return None



class HybridClosePriceProvider:
    """
    联网优先 + 股票池收盘价表兜底。

    取价顺序：
    1）先用 AKShare 联网取 A股/ETF/港股收盘价；
    2）联网接口没有数据、私募产品无法联网、或网络失败时，自动到“股票池2025年收盘价.xlsx”里找；
    3）两边都没有，才写入缺失价格。
    """
    def __init__(self, online_provider, fallback_prices=None, fallback_file=None, lookback_days=240, log_func=None):
        self.online = online_provider
        self.fallback_prices = fallback_prices or {}
        self.fallback_file = fallback_file
        self.lookback_days = int(lookback_days)
        self.log_func = log_func
        self.missing = []
        self.fallback_hit_count = 0

    def save_cache(self):
        if self.online:
            self.online.save_cache()

    def get_close_price(self, code, target_date, name='', max_lookback=None):
        lookback = int(max_lookback or self.lookback_days)
        before_missing = len(self.online.missing) if self.online else 0

        # 先联网取价。
        if self.online:
            try:
                v = self.online.get_close_price(code, target_date, name=name, max_lookback=lookback)
                if v is not None:
                    return v
            except Exception:
                v = None

        # 联网没有，再用股票池收盘价表兜底。
        v = find_local_close_price(self.fallback_prices, code, target_date, max_lookback=lookback)
        if v is not None:
            self.fallback_hit_count += 1
            # 联网 provider 可能已经把该条记为缺失；兜底成功后不应再显示在缺失 sheet。
            if self.online and len(self.online.missing) > before_missing:
                del self.online.missing[before_missing:]
            return v

        # 两边都没有，保留联网缺失原因；如果联网层没生成缺失，则补一条。
        if self.online and len(self.online.missing) > before_missing:
            self.missing.extend(self.online.missing[before_missing:])
        else:
            d = target_date.date() if isinstance(target_date, datetime) else target_date
            self.missing.append({
                'date': d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d),
                'code': normalize_sec_code(code),
                'name': str(name or ''),
                'market': 'LOCAL',
                'error': '联网无数据，股票池收盘价表也无数据',
            })
        return None

def get_fridays(start, end):
    """返回 [start, end] 区间内所有周五，首尾强制插入 start 和 end。

    注意：年度净值计算时，即使 12 月 31 日不是周五、甚至流水没有到 12 月 31 日，
    也必须把 12 月 31 日作为最后一个估值日。这个函数本身负责把 end 强制加入。
    """
    offset = (4 - start.weekday()) % 7
    d = start + timedelta(days=offset)
    fridays = []
    while d <= end:
        fridays.append(d)
        d += timedelta(days=7)
    if not fridays or fridays[0] != start:
        fridays.insert(0, start)
    if fridays[-1] != end:
        fridays.append(end)
    # 去重并保持顺序，避免 start/end 正好也是周五时重复。
    out = []
    seen = set()
    for x in fridays:
        if x not in seen:
            out.append(x)
            seen.add(x)
    return out


def force_year_end_date(nav_start, nav_end, transactions=None):
    """年度净值强制加入当年 12 月 31 日。

    有些券商流水导出只到最后一个交易日或最后一笔业务日，例如中信流水可能没有
    12 月 31 日；但会计/年度净值必须计算 12 月 31 日。
    因此只要是年度区间，就把结束估值日修正到该年度 12 月 31 日。
    
    若文件名本来就是 20250101_20251231，则不变；若文件名或流水只到
    20251226/20251230，也会补到 20251231。
    """
    year = nav_end.year if nav_end else nav_start.year
    # 如果交易流水年份更晚，以流水最大年份为准，避免跨年文件被截断。
    if transactions:
        try:
            year = max(year, max(t['date'].year for t in transactions if t.get('date')))
        except Exception:
            pass
    year_end = date(year, 12, 31)
    if nav_end < year_end:
        return year_end
    return nav_end


def get_hkd_rate(hkd_rates, d):
    """获取港股通汇率，找不到则向前最多回溯10天"""
    for delta in range(11):
        key = (d - timedelta(days=delta)).strftime('%Y-%m-%d')
        if key in hkd_rates:
            return hkd_rates[key]
    return 0.94


def find_close_price(price_provider, code, target_date, name='', max_lookback=240):
    """兼容旧函数名：从价格提供器中取收盘价。"""
    if hasattr(price_provider, 'get_close_price'):
        return price_provider.get_close_price(code, target_date, name=name, max_lookback=max_lookback)
    return None


def is_hk_security(code, name=''):
    """判断是否为港股。

    通达信/股票池里的港股价格是港币价，例如 HK02269/02269.HK/02269。
    估值时必须折算为人民币：港币收盘价 × 港币兑人民币汇率。
    """
    raw = str(code or '').strip().upper()
    nm = str(name or '').strip().upper()
    code_norm = normalize_sec_code(code)
    if raw.startswith('HK') or raw.endswith('.HK'):
        return True
    if '港' in nm or nm.endswith('H') or 'HK' in nm:
        return True
    # A股、ETF、北交所通常是6位；港股常见为5位，如02269、03968、02318、00700。
    if code_norm.isdigit() and len(code_norm) == 5:
        return True
    return False


def is_convertible_bond_code(code, name=''):
    """判断是否为可转债代码。"""
    code_s = normalize_sec_code(code)
    name_s = str(name or '')
    return (
        '转债' in name_s
        or (code_s.isdigit() and len(code_s) == 6 and code_s.startswith(('110', '111', '113', '118', '123', '127', '128')))
    )


def normalize_convertible_bond_close(close, code, name=''):
    """修正通达信转债价格口径。

    当前本地通达信表里转债可能显示为 1300.01，而交易流水成交均价为 130.001。
    对转债代码，若价格异常大于 500，则按除以 10 后的交易软件口径估值。
    """
    if close is None:
        return None
    try:
        v = float(close)
    except Exception:
        return close
    if is_convertible_bond_code(code, name) and v > 500:
        return v / 10.0
    return v


def record_missing_price(price_provider, code, target_date, name='', reason='通达信收盘价表和股票池收盘价表均无数据'):
    """在所有兜底逻辑都失败后，再记录缺失价格。

    之前版本是在价格表查不到时立即记录，导致“新股入账价/发行价兜底”已经成功估值，
    但仍然出现在“缺失价格”sheet。这里改为最终确认无估值价后才记录。
    """
    if not hasattr(price_provider, 'missing'):
        return
    d = target_date.date() if isinstance(target_date, datetime) else target_date
    item = {
        'date': d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d),
        'code': normalize_sec_code(code),
        'name': str(name or ''),
        'market': 'LOCAL',
        'error': reason,
    }
    k = (item['date'], item['code'], item['name'], item['market'])
    for old in price_provider.missing:
        if (old.get('date'), old.get('code'), old.get('name'), old.get('market')) == k:
            return
    price_provider.missing.append(item)

def calc_position_value(positions, price_provider, target_date, hkd_rates, cash, ipo_price=None):
    """计算账户总市值，返回 (总市值, 持仓明细)。

    本地收盘价缺失时，仍保留“新股入账价”兜底，避免新股尚未能取到价格时市值直接为 0。
    港股估值：通达信/股票池表中的港币收盘价 × 数量 × 港币兑人民币汇率。
    明细中同时保留原币收盘价、汇率、折人民币价格，方便核对。
    """
    hkd_rate = get_hkd_rate(hkd_rates, target_date)
    total = cash
    details = {}
    for code, (name, qty, cost) in positions.items():
        if qty == 0:
            continue
        close = find_close_price(price_provider, code, target_date, name=name)
        if close is None and ipo_price:
            close = ipo_price.get(normalize_sec_code(code)) or ipo_price.get(code)
        close = normalize_convertible_bond_close(close, code, name)
        # 配债缴款到正式转债上市之间，价格表通常没有转债行情。
        # 此时按缴款成本价估值，避免现金已扣但资产被计为 0。
        if close is None and cost and ('转债' in str(name) or '配债' in str(name)):
            close = float(cost)
        hk_flag = is_hk_security(code, name)
        if close is not None:
            currency = 'HKD' if hk_flag else 'CNY'
            fx_rate = hkd_rate if hk_flag else 1.0

            # 港股价格表中的收盘价是港币原币价。
            # 交易软件通常按 3 位小数展示港股价格，因此估值时也先把港币价保留 3 位，
            # 再乘以港币兑人民币汇率，人民币价保留 6 位，避免出现 52.799999 这类浮点尾数。
            if hk_flag:
                close_display = round(float(close), 3)
                close_rmb = round(close_display * float(fx_rate), 6)
            else:
                close_display = round(float(close), 6)
                close_rmb = round(close_display * float(fx_rate), 6)

            mv = round(close_rmb * qty, 2)
            total += mv
            details[code] = (name, qty, close_display, mv, currency, fx_rate, close_rmb)
        else:
            record_missing_price(price_provider, code, target_date, name=name)
            details[code] = (name, qty, None, 0, '', None, None)
    return round(total, 2), details


def compute_capital_weighted_sum(initial_total, transfers, target_date, start_date):
    """计算日期时间加权投入ofi"""
    total_days = max((target_date - start_date).days, 1)
    wsum = initial_total * total_days
    for tf in transfers:
        days_left = (target_date - tf['date']).days
        if days_left > 0:
            wsum += tf['amount'] * days_left
    return wsum, total_days


def plot_nav_chart(nav_records, save_path, account_name=''):
    """绘制净值折线图。

    图表标题使用当前处理的账户名称，而不是固定写死为某一个账户。
    account_name 优先来自期初/数据源文件名，例如“华泰雅兴250101_251231.xlsx” -> “华泰雅兴”。
    """
    dates     = [r[0] for r in nav_records]
    nav_vals  = [r[5] for r in nav_records]
    if dates:
        chart_year = dates[-1].year
    else:
        chart_year = datetime.now().year
    account_name = str(account_name or '').strip() or '账户'

    configure_matplotlib_chinese_font()

    fig, ax = plt.subplots(figsize=(14, 6))

    ax.plot(dates, nav_vals, color='#C00000', linewidth=2, label='单位净值', zorder=3)
    ax.fill_between(dates, 1.0, nav_vals,
                    where=[v >= 1.0 for v in nav_vals],
                    alpha=0.12, color='#C00000', label='超过1.0')
    ax.fill_between(dates, nav_vals, 1.0,
                    where=[v < 1.0 for v in nav_vals],
                    alpha=0.12, color='#006100', label='低于1.0')

    ax.axhline(y=1.0, color='#404040', linewidth=1, linestyle='--', alpha=0.6)

    max_idx = nav_vals.index(max(nav_vals))
    min_idx = nav_vals.index(min(nav_vals))
    ax.annotate(f'最高 {nav_vals[max_idx]:.4f}',
                xy=(dates[max_idx], nav_vals[max_idx]),
                xytext=(10, 8), textcoords='offset points',
                fontsize=9, color='#C00000',
                arrowprops=dict(arrowstyle='->', color='#C00000', lw=1))
    ax.annotate(f'最低 {nav_vals[min_idx]:.4f}',
                xy=(dates[min_idx], nav_vals[min_idx]),
                xytext=(10, -18), textcoords='offset points',
                fontsize=9, color='#006100',
                arrowprops=dict(arrowstyle='->', color='#006100', lw=1))

    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    ax.xaxis.set_major_locator(mdates.MonthLocator())
    plt.xticks(rotation=45, fontsize=9)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{v:.4f}'))
    ax.set_ylim(min(nav_vals) * 0.97, max(nav_vals) * 1.03)

    ax.set_title(f'{account_name}{chart_year} 单位净值走势', fontsize=14, fontweight='bold', pad=12)
    ax.set_xlabel('日期', fontsize=10)
    ax.set_ylabel('单位净值', fontsize=10)
    ax.legend(fontsize=9, loc='upper left')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    ax.grid(axis='x', linestyle=':', alpha=0.3)

    fig.tight_layout()
    fig.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close(fig)


def calc_vat():
    """
    计算股票卖出应缴增值税（资管产品/基金：转让金融商品差价收益按6%增值税）
    规则：
      - 计税依据：卖出收入 - 加权平均成本（买卖金额均按成交数量×成交均价）
      - 仅对正差价（盈利）征税；亏损可结转至下期，不退税
      - 税率 6%（一般纳税人），税额 = 差价/(1+6%) * 6%
      - 期末未抵扣的累计亏损不退税
      - 港股按人民币差价计税
    结果存为：原数据源文件名 + 应交税款.xlsx
    """
    text_pad.config(state=tk.NORMAL)
    text_pad.delete(1.0, tk.END)

    if state['book'] is None:
        messagebox.showerror('错误', '请先执行"一键处理"生成交易流水数据')
        return

    init_path = var_path_src.get().strip()
    if not init_path:
        messagebox.showerror('错误', '请先在"选择数据源文件"中选择期初持仓文件')
        return

    _base, _ext = os.path.splitext(init_path)
    output_path = _base + '应交税款' + _ext

    try:
        # ── 1. 读取期初成本价 → {code_str: cost_price} ──
        wb_init = openpyxl.load_workbook(init_path, data_only=True)
        ws_init = wb_init['期初']
        init_cost = {}   # {code: cost_price}
        init_qty  = {}   # {code: qty}
        for row in ws_init.iter_rows(min_row=2, values_only=True):
            code, name, qty, close, mkt, cost = (row[i] for i in range(6))
            if code is None or qty is None:
                continue
            code_s = str(code).strip().lstrip('0') or '0'
            # 标准化：A股6位，港股5位（补前导零）
            if code_s.isdigit():
                if len(code_s) <= 6 and len(str(code).strip()) == 6:
                    code_s = str(code).strip()
                elif len(code_s) <= 5 and len(str(code).strip()) == 5:
                    code_s = str(code).strip()
            c = float(cost) if cost else (float(close) if close else 0.0)
            init_cost[code_s] = c
            init_qty[code_s]  = int(qty) if qty else 0

        text_pad.insert('insert', f'期初成本表：{len(init_cost)} 只证券\n')
        text_pad.insert('insert', '买入成本口径：成交数量 × 成交均价（不使用发生金额）\n')
        text_pad.insert('insert', '买入/卖出金额口径：统一使用对账单“成交数量 × 成交均价”（不使用发生金额/成交金额）\n')

        # ── 2. 提取用于税款计算的交易流水 ──
        # 按用户确认口径：买入成本、卖出收入都统一使用“成交数量 × 成交均价”。
        # 因为部分券商只有“对账单”，没有“交割单/成交金额”；且“发生金额”会混入手续费、印花税等清算项。
        # 所以税款计算优先使用原始文件“对账单”，没有对账单时再退回当前已整理流水。
        def _load_vat_df_from_source():
            try:
                xls = pd.ExcelFile(init_path)
                for sheet in ['对账单', '交割单', '资金流水', '流水', '明细', '交易流水']:
                    if sheet in xls.sheet_names:
                        raw = pd.read_excel(init_path, sheet_name=sheet, dtype=str)
                        text_pad.insert('insert', f'税款计算数据源：原始文件“{sheet}”（买卖金额=成交数量×成交均价）\n')
                        return raw
            except Exception as e:
                text_pad.insert('insert', f'读取原始流水失败，改用已整理流水：{e}\n')
            text_pad.insert('insert', '税款计算数据源：已整理流水（买卖金额=成交数量×成交均价）\n')
            return state['book'].copy()

        df = _load_vat_df_from_source()

        # 标准化列名（兼容不同格式文件的列名差异）
        col_map = {}
        for c in df.columns:
            cs = str(c).strip()
            if cs in ('成交日期', '交收日期', '发生日期', '业务日期', '清算日期', '日期') or '日期' in cs:
                col_map[c] = '日期'
            elif cs in ('操作', '业务名称', '业务标志名称', '交易类别', '摘要', '业务摘要', '业务描述') or '业务' in cs:
                col_map[c] = '业务名称'
            elif '证券代码' in cs:
                col_map[c] = '证券代码'
            elif '证券名称' in cs:
                col_map[c] = '证券名称'
            elif '成交均价' in cs or '成交价格' in cs or cs in ('成交价', '价格'):
                col_map[c] = '成交均价'
            elif '成交数量' in cs or '成交股份' in cs or '发生数量' in cs or '股份发生数' in cs:
                col_map[c] = '成交数量'
            elif '成交金额' in cs:
                col_map[c] = '成交金额'
            elif '发生金额' in cs:
                col_map[c] = '发生金额'
        df = df.rename(columns=col_map)
        # 处理极端情况下重名列导致 df['列名'] 返回 DataFrame 的问题
        df = df.loc[:, ~df.columns.duplicated()].copy()

        required = ['日期', '业务名称', '证券代码', '成交均价', '成交数量']
        missing_cols = [c for c in required if c not in df.columns]
        if missing_cols:
            messagebox.showerror('错误', f'流水数据缺少列：{missing_cols}')
            return

        # ── 3. 标准化代码 & 日期 ──
        def norm_code(v):
            s = str(v).strip().upper().replace(' ', '')
            if s.endswith('.0'):
                s = s[:-2]
            s = re.sub(r'\.[A-Z]+$', '', s)
            s = re.sub(r'^(SH|SZ|BJ|NQ|HK)', '', s)
            # 这里不再 lstrip('0')，避免把港股 06030 变成 6030 后又被当 A股补成 006030。
            return s

        df['_code'] = df['证券代码'].apply(norm_code)
        # 日期可能是整数（如20250102）或datetime，统一转datetime并覆盖原列
        raw_dates = df['日期']
        if raw_dates.dtype in ('int64', 'float64', 'int32', 'float32'):
            df['_date'] = pd.to_datetime(raw_dates.astype(str), format='%Y%m%d', errors='coerce')
        else:
            df['_date'] = pd.to_datetime(raw_dates, errors='coerce')
        df = df.dropna(subset=['_date']).sort_values('_date').reset_index(drop=True)
        # 将日期列替换为正确datetime，确保输出时格式正确
        df['日期'] = df['_date']

        formal_bond_by_root = {}
        for _, r in df.iterrows():
            nm = str(r.get('证券名称', '') or '')
            if '转债' in nm and str(r.get('_code', '') or ''):
                formal_bond_by_root[_convertible_root(nm)] = str(r.get('_code')).strip()

        # ── 4. 过滤有效买卖操作 ──
        def is_buy(biz):
            b = str(biz)
            return '买入' in b or '新股入账' in b or '新股入帐' in b or '配股' in b or '配债缴款' in b

        def is_sell(biz):
            b = str(biz)
            return '卖出' in b

        # ── 5. 建立持仓成本池（加权平均法）──
        # cost_pool: {code: {'qty': int, 'total_cost': float}}
        cost_pool = {}
        for code, cost in init_cost.items():
            qty = init_qty.get(code, 0)
            cost_pool[code] = {'qty': qty, 'total_cost': cost * qty}

        VAT_RATE = 0.06   # 增值税率 6%

        # ── 6. 逐笔处理，计算应缴税款 ──
        vat_col      = [0.0] * len(df)
        hold_qty_col = [''] * len(df)   # 本笔操作后持有数量
        cost_px_col  = [''] * len(df)   # 本笔操作后加权成本价
        formula_col  = [''] * len(df)   # 税款计算公式说明
        carry_loss   = {}               # 按代码分别结转亏损 {code: negative_amount}

        for idx, row in df.iterrows():
            biz  = str(row['业务名称']) if pd.notna(row['业务名称']) else ''
            code = row['_code']
            if '配债' in str(row.get('证券名称', '') or ''):
                code = formal_bond_by_root.get(_convertible_root(row.get('证券名称', '')), code)
            qty_raw = row['成交数量'] if pd.notna(row['成交数量']) else 0
            price   = float(row['成交均价']) if pd.notna(row['成交均价']) else 0.0
            amt_raw = float(row['发生金额']) if '发生金额' in df.columns and pd.notna(row['发生金额']) else None
            trade_amt_raw = float(row['成交金额']) if '成交金额' in df.columns and pd.notna(row['成交金额']) else None

            qty = abs(int(qty_raw)) if qty_raw != 0 else 0

            if is_buy(biz) and qty > 0 and price > 0:
                # 买入：更新成本池（加权平均）
                # 按用户确认口径：买入成本统一使用“成交数量 × 成交均价”。
                # 不再使用发生金额/成交金额，避免把含其他清算项、费用或港股通差异的金额带入买入成本。
                buy_cost = price * qty
                if code not in cost_pool:
                    cost_pool[code] = {'qty': 0, 'total_cost': 0.0}
                cost_pool[code]['qty']        += qty
                cost_pool[code]['total_cost'] += buy_cost
                # 记录买入后持仓情况
                pool_after = cost_pool[code]
                new_avg = pool_after['total_cost'] / pool_after['qty'] if pool_after['qty'] > 0 else 0
                hold_qty_col[idx] = pool_after['qty']
                cost_px_col[idx]  = round(new_avg, 4)
                formula_col[idx]  = (
                    f"买入{qty}股，买入成本{abs(buy_cost):.2f}，"
                    f"新加权均价=总成本{pool_after['total_cost']:.2f}÷持仓{pool_after['qty']}股"
                    f"={new_avg:.4f}"
                )

            elif is_sell(biz) and qty > 0 and price > 0:
                # 卖出：计算差价税
                pool = cost_pool.get(code, {'qty': 0, 'total_cost': 0.0})
                if pool['qty'] > 0:
                    avg_cost = pool['total_cost'] / pool['qty']
                else:
                    avg_cost = init_cost.get(code, price)

                # 卖出收入：按用户确认口径，统一使用“成交数量 × 成交均价”。
                # 不使用发生金额，也不使用成交金额，保证与只有对账单的券商口径一致。
                sell_income = price * qty
                sell_income_label = '成交数量×成交均价'
                sell_cost   = avg_cost * qty
                diff_before_carry = sell_income - sell_cost
                # 按代码取前期结转亏损
                prev_carry = carry_loss.get(code, 0.0)
                diff = diff_before_carry + prev_carry

                if diff > 0:
                    vat = diff / (1 + VAT_RATE) * VAT_RATE
                    vat_col[idx] = round(vat, 2)
                    # 盈利抵扣掉前期亏损后，清零该代码的结转亏损
                    carry_loss[code] = 0.0
                    carry_desc = f"，抵扣前期亏损{prev_carry:.2f}" if prev_carry != 0 else ""
                    formula_col[idx] = (
                        f"卖出收入({sell_income_label}){sell_income:.2f}-成本({avg_cost:.4f}×{qty}={sell_cost:.2f})"
                        f"{carry_desc}=差价{diff:.2f}；"
                        f"税额={diff:.2f}÷1.06×6%={round(vat,2):.2f}"
                    )
                else:
                    # 亏损：按代码累计结转
                    carry_loss[code] = diff
                    formula_col[idx] = (
                        f"卖出收入({sell_income_label}){sell_income:.2f}-成本({avg_cost:.4f}×{qty}={sell_cost:.2f})"
                        f"=差价{diff_before_carry:.2f}，累计亏损{diff:.2f}结转下期，本笔不缴税"
                    )

                # 更新成本池（减少持仓）
                if pool['qty'] >= qty:
                    cost_pool[code]['qty']        -= qty
                    cost_pool[code]['total_cost'] -= avg_cost * qty
                else:
                    cost_pool[code] = {'qty': 0, 'total_cost': 0.0}

                # 记录卖出后持仓情况
                pool_after = cost_pool.get(code, {'qty': 0, 'total_cost': 0.0})
                new_avg2 = (pool_after['total_cost'] / pool_after['qty']
                            if pool_after['qty'] > 0 else avg_cost)
                hold_qty_col[idx] = pool_after['qty']
                cost_px_col[idx]  = round(new_avg2, 4) if pool_after['qty'] > 0 else ''

            else:
                # 非买卖操作：如果该代码在成本池有持仓，填当前持仓快照
                if code in cost_pool and cost_pool[code]['qty'] > 0:
                    p = cost_pool[code]
                    hold_qty_col[idx] = p['qty']
                    cost_px_col[idx]  = round(p['total_cost'] / p['qty'], 4)

        # ── 7. 将新列插入 df ──
        # 找到"应缴增值税"准备插入的位置：在原始列末尾依次插入
        df['持有数量']   = hold_qty_col
        df['成本价']     = cost_px_col
        df['应缴增值税'] = vat_col
        df['税款计算公式'] = formula_col

        # ── 8. 汇总统计 ──
        total_vat    = df['应缴增值税'].sum()
        sell_rows    = df[df['应缴增值税'] > 0]
        text_pad.insert('insert', f'卖出笔数（含税）：{len(sell_rows)} 笔\n')
        text_pad.insert('insert', f'累计应缴增值税：{total_vat:,.2f} 元\n')
        total_carry = sum(carry_loss.values())
        text_pad.insert('insert', f'期末累计亏损结转：{total_carry:,.2f} 元\n')
        text_pad.insert('insert', f'增值税率：{VAT_RATE*100:.0f}%（一般纳税人）\n\n')
        text_pad.insert('insert', f'正在保存 → {output_path}\n')

        # ── 9. 保存结果 ──
        # 还原列名、去掉内部辅助列
        save_df = df.drop(columns=['_code', '_date'], errors='ignore')
        # 日期列转为字符串避免格式问题
        for dc in save_df.columns:
            if 'date' in str(dc).lower() or dc == '日期':
                try:
                    col = save_df[dc]
                    # 整数列（如20250102）需要先转str再按格式解析
                    if pd.api.types.is_integer_dtype(col):
                        save_df[dc] = pd.to_datetime(col.astype(str), format='%Y%m%d').dt.strftime('%Y-%m-%d')
                    else:
                        save_df[dc] = pd.to_datetime(col).dt.strftime('%Y-%m-%d')
                except Exception:
                    pass

        # ── 调整列顺序：原始列 + 持有数量 + 成本价 + 应缴增值税 + 税款计算公式 ──
        orig_cols = [c for c in save_df.columns
                     if c not in ('持有数量', '成本价', '应缴增值税', '税款计算公式')]
        save_df = save_df[orig_cols + ['持有数量', '成本价', '应缴增值税', '税款计算公式']]

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            save_df.to_excel(writer, index=False, sheet_name='应交税款')
            ws_out = writer.sheets['应交税款']
            from openpyxl.styles import Alignment
            ncols = len(save_df.columns)
            col_names = list(save_df.columns)

            # 各列格式
            idx_qty    = col_names.index('持有数量')     + 1
            idx_cost   = col_names.index('成本价')       + 1
            idx_vat    = col_names.index('应缴增值税')   + 1
            idx_fml    = col_names.index('税款计算公式') + 1

            for r in range(2, len(save_df) + 2):
                ws_out.cell(r, idx_qty).number_format  = '#,##0'
                ws_out.cell(r, idx_cost).number_format = '#,##0.0000'
                ws_out.cell(r, idx_vat).number_format  = '#,##0.00'
                ws_out.cell(r, idx_fml).alignment      = Alignment(wrap_text=False)

            # 税款公式列列宽加宽
            ws_out.column_dimensions[
                ws_out.cell(1, idx_fml).column_letter
            ].width = 80

        text_pad.insert('insert', f'完成！共 {len(save_df)} 行，应缴增值税列已写入\n')
        update_status(f'应纳税款计算完成，共 {total_vat:,.2f} 元')
        messagebox.showinfo('完成', f'应缴增值税汇总：{total_vat:,.2f} 元\n\n已保存至：\n{output_path}')

    except Exception as e:
        import traceback
        text_pad.insert('insert', f'错误：{e}\n{traceback.format_exc()}\n')
        messagebox.showerror('错误', str(e))

    text_pad.config(state=tk.DISABLED)


def compute_nav():
    """主净值计算函数，使用 state['book']（一键处理后的 DataFrame）作为交易流水来源"""
    text_pad.config(state=tk.NORMAL)
    text_pad.delete(1.0, tk.END)

    if state['book'] is None:
        messagebox.showerror('错误', '请先执行"一键处理"生成交易流水数据')
        return

    # ── 动态路径：从"选择数据源文件"输入框派生 ──
    init_path = var_path_src.get().strip()
    if not init_path:
        messagebox.showerror('错误', '请先在"选择数据源文件"中选择期初持仓文件')
        return
    _base, _ext = os.path.splitext(init_path)
    output_path = _base + '净值' + _ext      # 如：万联正涵2025净值.xlsx
    chart_path  = _base + '净值图.png'        # 如：万联正涵2025净值图.png

    text_pad.insert('insert', '正在加载数据...\n\n')

    try:
        logic_path = resolve_existing_file(
            LOGIC_FILE,
            base_file=init_path,
            candidates=['交易流水业务变化逻辑.xlsx']
        )
        tdx_price_path = resolve_existing_file(
            TDX_PRICE_FILE,
            base_file=init_path,
            candidates=['通达信提取收盘价结果.xlsx', 'tdx_p_close.xlsx', 'tdxprice.xlsx']
        )
        fallback_price_path = resolve_existing_file(
            PRICE_FILE,
            base_file=init_path,
            candidates=['股票池2025年收盘价.xlsx']
        )

        rules                              = load_business_rules(logic_path)
        positions, cash, hkd_rates, init   = load_initial_positions(init_path)
        transactions                       = load_transactions_from_df(state['book'])
        convertible_alias                  = build_convertible_bond_alias_map(transactions)
        bank_transfers                     = load_bank_transfers(transactions)
        tdx_prices                         = load_stock_pool_close_prices(tdx_price_path)
        fallback_prices                    = load_stock_pool_close_prices(fallback_price_path)
        prices                             = LocalPriorityClosePriceProvider(
            primary_prices=tdx_prices,
            fallback_prices=fallback_prices,
            primary_file=tdx_price_path,
            fallback_file=fallback_price_path,
            lookback_days=240,
            log_func=lambda s: text_pad.insert('insert', s + '\n')
        )

        if len(transactions) == 0:
            messagebox.showerror('错误', '交易流水为空，请检查一键处理后的数据')
            return

        text_pad.insert('insert', f'  业务规则: {len(rules)} 条\n')
        text_pad.insert('insert', f'  期初持仓: {len(positions)} 只, 现金: {cash:,.2f}, 期初总资产: {init:,.2f}\n')
        text_pad.insert('insert', f'  交易流水: {len(transactions)} 条\n')
        if convertible_alias:
            text_pad.insert('insert', f'  配债转债映射: {len(convertible_alias)} 组\n')
        text_pad.insert('insert', f'  银行转账: {len(bank_transfers)} 笔\n')
        text_pad.insert('insert', f'  收盘价:   通达信收盘价表优先，股票池收盘价表兜底，不联网；港股按汇率折人民币估值\n')
        text_pad.insert('insert', f'             通达信表：{tdx_price_path}，已读取 {len(tdx_prices)} 条价格/净值\n')
        text_pad.insert('insert', f'             股票池表：{fallback_price_path}，已读取 {len(fallback_prices)} 条价格/净值\n')

        first_date = transactions[0]['date']
        last_date  = transactions[-1]['date']
        # 从期初文件名解析日期区间（格式：万联正涵YYYYMMDD_YYYYMMDD.xlsx）
        fname = os.path.splitext(os.path.basename(init_path))[0]
        m = re.search(r'(\d{8})_(\d{8})$', fname)
        if m:
            nav_start = datetime.strptime(m.group(1), '%Y%m%d').date()
            nav_end   = datetime.strptime(m.group(2), '%Y%m%d').date()
        else:
            nav_start = first_date
            nav_end   = last_date

        # 年度净值必须包含 12 月 31 日：
        # 有些券商流水没有导出到 12/31，且 12/31 不一定是周五，
        # 但作为一年最后一天仍需按当日/向前回溯价格计算净值。
        original_nav_end = nav_end
        nav_end = force_year_end_date(nav_start, nav_end, transactions)
        fridays    = get_fridays(nav_start, nav_end)
        capital_events = build_capital_events_from_bank_transfers(init, bank_transfers, nav_start)
        text_pad.insert('insert', f'  流水日期范围: {first_date} ~ {last_date}\n')
        if nav_end != original_nav_end:
            text_pad.insert('insert', f'  估值结束日: {original_nav_end} → {nav_end}（已强制加入年度最后一天 12-31）\n')
        text_pad.insert('insert', f'  估值日期: 周五 + 首日 + 期末日，共 {len(fridays)} 个\n')
        text_pad.insert('insert', f'  投入积数: 不使用证券账户持仓表；按期初总资产 + 流水银行转存/转取计算；仅新股托管转入按数量增加持仓，资金事件 {len(capital_events)} 笔\n\n')

        # 预构建"新股入账"价格字典：{sec_code: price}（用于收盘价缺失时的兜底）
        ipo_price = {}
        for txn in transactions:
            if txn['biz_name'] == '新股入账' and txn['sec_code'] and txn['price'] and txn['price'] > 0:
                ipo_price[normalize_sec_code(txn['sec_code'])] = txn['price']

        # 只有新股相关的托管转入才影响持仓。
        # 普通托管转入/转托管不增减本账户实际持仓。
        ipo_codes = build_ipo_code_set(transactions)

        txn_by_date  = defaultdict(list)
        for txn in transactions:
            txn_by_date[txn['date']].append(txn)
        trading_days = sorted(txn_by_date)
        day_idx      = 0

        nav_records = []
        text_pad.insert('insert', '开始逐日模拟持仓变动...\n')

        for friday in fridays:
            while day_idx < len(trading_days) and trading_days[day_idx] <= friday:
                for txn in txn_by_date[trading_days[day_idx]]:
                    biz, code, name, qty, amount = (
                        txn['biz_name'], txn['sec_code'], txn['sec_name'],
                        txn['qty'], txn['amount']
                    )
                    if is_convertible_allotment_payment(txn) and code:
                        target_code = convertible_alias.get(code, code)
                        target_name = re.sub(r'配债$', '转债', str(name or target_code))
                        unit_cost = float(amount or 0.0) / qty if qty else 0.0
                        if target_code in positions:
                            n, q, c = positions[target_code]
                            positions[target_code] = (target_name or n, q + qty, c or unit_cost)
                        else:
                            positions[target_code] = (target_name, qty, unit_cost)
                        cash -= float(amount or 0.0)
                        continue
                    if (
                        is_convertible_allotment_intermediate(txn)
                        or (code in convertible_alias and biz == '权证上账')
                        or (code in set(convertible_alias.values()) and biz in ('股份上市', '交收股份修正', '交收股份修正取消'))
                    ):
                        continue
                    rule = rules.get(biz)
                    if rule is None:
                        continue
                    qty_rule, amt_rule = rule

                    # 只有“新股相关托管转入/限售股份转无限售”才按流水数量增加持仓。
                    # 普通托管转入、转托管不改变本账户真实持仓，仍按业务规则表处理为“不变”。
                    # 新股判断依据：备注/摘要/操作包含“新股”或“XGJX”，或该代码曾出现新股入账/申购中签。
                    if biz == '托管转入' and qty > 0 and is_new_stock_custody_transfer(txn, ipo_codes):
                        qty_rule = '加'

                    if qty_rule == '加' and code:
                        if code in positions:
                            n, q, c = positions[code]
                            positions[code] = (name or n, q + qty, c)
                        else:
                            positions[code] = (name or code, qty, 0.0)
                    elif qty_rule == '减' and code:
                        if code in positions:
                            n, q, c = positions[code]
                            new_q = q - qty
                            if new_q <= 0:
                                del positions[code]
                            else:
                                positions[code] = (n, new_q, c)

                    if   amt_rule == '加': cash += amount
                    elif amt_rule == '减': cash -= amount
                day_idx += 1

            total_value, details = calc_position_value(
                positions, prices, friday, hkd_rates, cash, ipo_price
            )

            # 按流水中的银行转存/转取计算资金投入，不读取证券账户持仓表。
            # 投入积数 = Σ资金净投入 × (估值日 - 操作日 + 1)；平均投入 = 投入积数 / 期间天数。
            cap_sum, total_days, avg_cap, net_capital = compute_capital_from_events(
                capital_events, friday, nav_start
            )
            profit = total_value - net_capital
            # 净值 = 1 + 当前盈亏 / 平均投入资金（平均投入即资金占用）
            nav_val = 1.0 + profit / avg_cap if avg_cap != 0 else 1.0

            nav_records.append((friday, total_value, details, cash, cap_sum, nav_val, avg_cap, net_capital, profit))
            pos_cnt = sum(1 for _, (_, q, _) in positions.items() if q > 0)
            text_pad.insert('insert',
                f'  {friday}  总市值:{total_value:>15,.2f}  净值:{nav_val:.4f}  '
                f'(持仓{pos_cnt}只, 现金{cash:,.2f})\n')

        # ── 写入 Excel ──
        text_pad.insert('insert', f'\n正在写入: {output_path}\n')
        wb_out = openpyxl.Workbook()
        ws     = wb_out.active
        ws.title = '净值明细'

        hf   = openpyxl.styles.Font(bold=True, size=11)
        df   = openpyxl.styles.Font(bold=True, size=11, color='1F4E79')
        sf   = openpyxl.styles.Font(bold=True, size=10)
        tf_r = openpyxl.styles.Font(bold=True, size=11, color='C00000')
        nf   = openpyxl.styles.Font(bold=True, size=11, color='006100')
        bdr  = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin', color='D9D9D9'))

        for ci, h in enumerate(['证券代码','证券名称','数量','收盘价(原币)','币种','估值汇率','人民币价','市值','投入积数(亿×天)','净值'], 1):
            ws.cell(row=1, column=ci, value=h).font = hf

        row_num = 2
        for rec in nav_records:
            friday, total_value, details, cash_val, cap_sum, nav_val, avg_cap = rec[:7]
            net_capital = rec[7] if len(rec) > 7 else None
            profit = rec[8] if len(rec) > 8 else None
            c = ws.cell(row=row_num, column=1, value=friday)
            c.number_format = 'YYYY-MM-DD'; c.font = df
            ws.cell(row=row_num, column=2, value=f'总资产: {total_value:,.2f}').font = tf_r
            # 投入积数单位：亿×天，缩小显示
            cap_y = cap_sum / 1e8
            ws.cell(row=row_num, column=9, value=f'投入积数: {cap_y:.2f} 亿×天').font = sf
            ws.cell(row=row_num, column=10, value=f'净值: {nav_val:.4f}').font = nf
            row_num += 1

            items = sorted(
                [(c, n, q, cl, mv, cur, fx, cl_rmb) for c, (n, q, cl, mv, cur, fx, cl_rmb) in details.items() if q > 0],
                key=lambda x: x[4], reverse=True
            )
            holding_sum = sum(mv for *_, mv, cur, fx, cl_rmb in items)

            for code, name, qty, close, mv, currency, fx_rate, close_rmb in items:
                ws.cell(row=row_num, column=1, value=code)
                ws.cell(row=row_num, column=2, value=name)
                c3 = ws.cell(row=row_num, column=3, value=qty)
                c3.number_format = '#,##0'
                # 收盘价(原币)：港股这里是港币价格，A股/基金为人民币价格。
                display_close = close
                if display_close is None:
                    display_close = ipo_price.get(normalize_sec_code(code)) or ipo_price.get(code)   # None 或 新股入账价
                c4 = ws.cell(row=row_num, column=4,
                             value=display_close if display_close is not None else '无数据')
                if display_close is not None:
                    # 港股原币收盘价按交易软件口径显示 3 位小数；A股/基金/私募仍保留 6 位。
                    c4.number_format = '#,##0.000' if currency == 'HKD' else '#,##0.000000'
                ws.cell(row=row_num, column=5, value=currency)
                c6 = ws.cell(row=row_num, column=6, value=fx_rate if fx_rate is not None else '')
                if fx_rate is not None:
                    c6.number_format = '#,##0.000000'
                c7 = ws.cell(row=row_num, column=7, value=close_rmb if close_rmb is not None else '')
                if close_rmb is not None:
                    c7.number_format = '#,##0.000000'
                c8 = ws.cell(row=row_num, column=8, value=round(mv, 2))
                c8.number_format = '#,##0.00'
                for ci in range(1, 11):
                    ws.cell(row=row_num, column=ci).border = bdr
                row_num += 1

            for label, val, font in [
                ('持仓市值', holding_sum, sf),
                ('现金',     cash_val,   sf),
                ('总资产',   total_value, tf_r),
                ('累计净投入', net_capital if net_capital is not None else 0, sf),
                ('平均投入', avg_cap,    sf),
                ('利润', profit if profit is not None else 0,    sf),
                ('净值',     nav_val,    nf),
            ]:
                ws.cell(row=row_num, column=2, value=label).font = font
                c8 = ws.cell(row=row_num, column=8, value=round(val, 2))
                c8.font = font
                c8.number_format = '0.0000' if label == '净值' else '#,##0.00'
                row_num += 1

            row_num += 1

        for col, w in zip('ABCDEFGHIJ', [14, 20, 14, 16, 10, 12, 16, 18, 18, 12]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A2'

        if hasattr(prices, 'save_cache'):
            prices.save_cache()
            if hasattr(prices, 'primary_hit_count'):
                text_pad.insert('insert', f"\n通达信收盘价表命中：{prices.primary_hit_count} 次\n")
            if hasattr(prices, 'fallback_hit_count') and prices.fallback_hit_count:
                text_pad.insert('insert', f"股票池收盘价表兜底命中：{prices.fallback_hit_count} 次\n")
            if getattr(prices, 'missing', None):
                uniq_missing = []
                seen_missing = set()
                for item in prices.missing:
                    k = (item.get('date'), item.get('code'), item.get('name'), item.get('market'))
                    if k in seen_missing:
                        continue
                    seen_missing.add(k)
                    uniq_missing.append(item)
                text_pad.insert('insert', f'\n本地收盘价缺失明细：{len(uniq_missing)} 条（已写入 Excel 的“缺失价格”sheet）\n')
                ws_miss = wb_out.create_sheet('缺失价格')
                ws_miss.append(['日期', '证券代码', '证券名称', '市场', '原因'])
                for item in uniq_missing:
                    ws_miss.append([item.get('date'), item.get('code'), item.get('name'), item.get('market'), item.get('error')])
                for col in range(1, 6):
                    ws_miss.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24

        # ── 资金投入明细：直接来自交易流水中的银行转存/转取 ──
        try:
            ws_cap = wb_out.create_sheet('资金投入明细')
            ws_cap.append(['操作日期', '业务名称', '资金净投入', '来源'])
            for ev in capital_events:
                ws_cap.append([ev.get('date'), ev.get('biz_name'), ev.get('amount'), ev.get('source')])
            for row in ws_cap.iter_rows(min_row=2, max_col=4):
                row[0].number_format = 'YYYY-MM-DD'
                row[2].number_format = '#,##0.00'
            for col, width in zip('ABCD', [14, 18, 18, 28]):
                ws_cap.column_dimensions[col].width = width
        except Exception as e:
            text_pad.insert('insert', f'  [提示] 资金投入明细 sheet 写入失败（{e}）\n')

        wb_out.save(output_path)

        # ── 绘制净值图 ──
        # 图表标题跟随当前账户名称，避免固定显示“万联正涵”。
        account_name_for_chart = _make_account_hint(init_path) or _make_account_hint(state.get('path_src', '')) or var_broker.get()
        plot_nav_chart(nav_records, chart_path, account_name_for_chart)

        # ── 嵌入图片到 Excel ──
        try:
            from openpyxl.drawing.image import Image as XLImage
            wb2 = openpyxl.load_workbook(output_path)
            ws_chart = wb2.create_sheet('净值走势图')
            img = XLImage(chart_path)
            img.anchor = 'A1'
            ws_chart.add_image(img)
            wb2.save(output_path)
        except Exception as e:
            text_pad.insert('insert', f'  [提示] 图片嵌入 Excel 失败（{e}），PNG 文件仍可独立查看\n')

        text_pad.insert('insert', f'\n完成！共 {len(nav_records)} 个估值日净值记录（含年度最后一天 12-31）\n')
        text_pad.insert('insert', f'Excel: {output_path}\n')
        text_pad.insert('insert', f'图表:  {chart_path}\n')
        update_status(f'✅ 净值计算完成，共 {len(nav_records)} 个估值日净值，已生成 Excel + 图表')

    except Exception as e:
        messagebox.showerror('净值计算错误', f'计算失败：{e}')
        text_pad.insert('insert', f'\n❌ 计算失败：{e}\n')


# ══════════════════════════════════════════════
# 六、自定义控件工厂
# ══════════════════════════════════════════════

def make_action_button(parent, text, command, **grid_kw):
    btn = tk.Button(
        parent, text=text, font=FONTS['button_bold'],
        bg=COLORS['action_bg'], fg=COLORS['action_fg'],
        activebackground=COLORS['action_hover'], activeforeground=COLORS['action_fg'],
        relief=tk.FLAT, cursor='hand2', bd=0, padx=12, pady=5, command=command,
    )
    btn.grid(**grid_kw)
    return btn


def make_secondary_button(parent, text, command, **grid_kw):
    btn = tk.Button(
        parent, text=text, font=FONTS['button'],
        bg=COLORS['secondary_bg'], fg=COLORS['secondary_fg'],
        activebackground=COLORS['secondary_hover'], activeforeground=COLORS['secondary_fg'],
        relief=tk.FLAT, cursor='hand2', bd=0, padx=12, pady=5, command=command,
    )
    btn.grid(**grid_kw)
    return btn


def make_open_button(parent, text, command, **grid_kw):
    btn = tk.Button(
        parent, text=text, font=FONTS['button'],
        bg=COLORS['open_bg'], fg=COLORS['open_fg'],
        activebackground=COLORS['open_hover'], activeforeground=COLORS['open_fg'],
        relief=tk.FLAT, cursor='hand2', bd=0, padx=12, pady=5, command=command,
    )
    btn.grid(**grid_kw)
    return btn


def make_step_row(parent, row, step_num, step_text, entry_var=None, browse_cmd=None):
    pad = {'padx': (10, 5), 'pady': (4, 4)}

    step_label = tk.Label(
        parent, text=f' {step_num} ',
        font=('Consolas', 11, 'bold'), bg=COLORS['accent'], fg='white',
        relief=tk.FLAT, padx=2, pady=0,
    )
    step_label.grid(row=row, column=0, padx=(10, 6), pady=pad['pady'], sticky=tk.W)

    text_label = tk.Label(
        parent, text=step_text, font=FONTS['step'],
        bg=COLORS['card_bg'], fg=COLORS['step_label'], anchor=tk.W,
    )
    text_label.grid(row=row, column=1, padx=(0, 8), pady=pad['pady'], sticky=tk.W)

    col_offset = 2

    if browse_cmd is not None:
        browse_btn = tk.Button(
            parent, text='📁 浏览', font=FONTS['button'],
            bg=COLORS['accent_light'], fg=COLORS['accent'],
            activebackground='#D6EAF8', activeforeground=COLORS['accent'],
            relief=tk.FLAT, cursor='hand2', bd=0, padx=10, pady=4, command=browse_cmd,
        )
        browse_btn.grid(row=row, column=col_offset, padx=(0, 6), pady=pad['pady'], sticky=tk.W)
        col_offset += 1

    if entry_var is not None:
        entry = ttk.Entry(parent, width=65, textvariable=entry_var, font=FONTS['entry'])
        entry.grid(row=row, column=col_offset, padx=(0, 10), pady=pad['pady'], sticky=tk.EW)
        return entry


def make_action_row(parent, row, action_text, action_cmd, save_text, save_cmd,
                    save_var, open_text, open_cmd, vat_text=None, vat_cmd=None):
    pad_y = (6, 8)

    make_action_button(parent, f'▶ {action_text}', action_cmd,
                       row=row, column=0, columnspan=2,
                       padx=(10, 8), pady=pad_y, sticky=tk.W)

    make_secondary_button(parent, f'💾 {save_text}', save_cmd,
                         row=row, column=2, padx=(8, 4), pady=pad_y, sticky=tk.W)

    save_entry = ttk.Entry(parent, width=28, textvariable=save_var, font=FONTS['entry'])
    save_entry.grid(row=row, column=3, padx=4, pady=pad_y, sticky=tk.W)

    if vat_text:
        make_secondary_button(parent, f'🧾 {vat_text}', vat_cmd,
                              row=row, column=4, padx=(4, 4), pady=pad_y, sticky=tk.W)
        open_col = 5
    else:
        open_col = 4

    make_open_button(parent, f'📂 {open_text}', open_cmd,
                     row=row, column=open_col, padx=(4, 10), pady=pad_y, sticky=tk.E)


def make_card_frame(parent, title, row, **grid_kw):
    outer = tk.Frame(parent, bg=COLORS['frame_border'], bd=1, relief=tk.FLAT)
    outer.grid(row=row, padx=12, pady=(0, 6), sticky=tk.NSEW, **grid_kw)
    outer.grid_columnconfigure(0, weight=1)

    title_bar = tk.Frame(outer, bg=COLORS['title_bg'], height=30)
    title_bar.grid(row=0, column=0, sticky=tk.NSEW)
    title_bar.grid_propagate(False)

    title_label = tk.Label(
        title_bar, text=f'  {title}', font=FONTS['subtitle'],
        bg=COLORS['title_bg'], fg=COLORS['title_fg'], anchor=tk.W,
    )
    title_label.pack(side=tk.LEFT, padx=(8, 0), fill=tk.Y, expand=True)

    content = tk.Frame(outer, bg=COLORS['card_bg'])
    content.grid(row=1, column=0, sticky=tk.NSEW)
    content.grid_columnconfigure(3, weight=1)
    outer.grid_rowconfigure(1, weight=1)

    return content


# ══════════════════════════════════════════════
# 七、界面构建
# ══════════════════════════════════════════════

root = tk.Tk()
root.title('券商数据处理程序')
root.geometry('960x720+120+60')
root.configure(bg=COLORS['bg'])
root.minsize(860, 620)
root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# ── 顶部标题栏 ──
header = tk.Frame(root, bg=COLORS['accent'], height=42)
header.grid(row=0, column=0, sticky=tk.NSEW)
header.grid_propagate(False)

tk.Label(
    header, text='📊  券商数据处理程序',
    font=FONTS['title'], bg=COLORS['accent'], fg='white',
).pack(side=tk.LEFT, padx=16, fill=tk.Y, expand=True)

tk.Label(
    header, text='支持券商：广发 | 国君 | 华泰 | 建投 | 万联 | 爱建 | 中信 | 长江 | 申万',
    font=FONTS['hint'], bg=COLORS['accent'], fg='#BEE3F8',
).pack(side=tk.RIGHT, padx=16, fill=tk.Y)

# ── 状态栏 ──
status_bar = tk.Frame(root, bg='#EDF2F7', height=24)
status_bar.grid(row=10, column=0, sticky=tk.NSEW)
status_bar.grid_propagate(False)

status_label = tk.Label(
    status_bar, text='就绪 — 请选择券商并导入数据文件',
    font=FONTS['hint'], bg='#EDF2F7', fg='#718096', anchor=tk.W,
)
status_label.pack(side=tk.LEFT, padx=12, fill=tk.Y)

tk.Label(
    status_bar, text='v3.1', font=('Consolas', 8),
    bg='#EDF2F7', fg='#A0AEC0',
).pack(side=tk.RIGHT, padx=12, fill=tk.Y)


def update_status(msg: str):
    status_label.config(text=msg)


# ── 主滚动区域 ──
main_frame = tk.Frame(root, bg=COLORS['bg'])
main_frame.grid(row=1, column=0, rowspan=9, sticky=tk.NSEW, padx=0, pady=8)
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_rowconfigure(2, weight=1)

# ── 界面变量 ──
var_path_src    = tk.StringVar()
var_path_fmt    = tk.StringVar()
var_path_induct = tk.StringVar()
var_path_merge  = tk.StringVar()
var_save        = tk.StringVar()
var_save1       = tk.StringVar()
var_save2       = tk.StringVar()
var_broker      = tk.StringVar()

# ────────────────────────────────────
# 面板 1：数据整理
# ────────────────────────────────────
panel1 = make_card_frame(main_frame, '数据整理', 0)

make_step_row(panel1, row=0, step_num='1', step_text='选择券商')

broker_box = ttk.Combobox(panel1, width=18, textvariable=var_broker,
                          values=list(BROKER_CONFIG.keys()),
                          font=FONTS['button'], state='readonly')
broker_box.grid(row=0, column=2, padx=(0, 6), pady=(4, 4), sticky=tk.W)
broker_box.bind('<<ComboboxSelected>>', lambda e: update_status(
    f'已选择券商：{var_broker.get()}'))

make_step_row(panel1, row=1, step_num='2', step_text='选择格式文件',
              entry_var=var_path_fmt, browse_cmd=select_fmt_file)

make_step_row(panel1, row=2, step_num='3', step_text='选择数据源文件',
              entry_var=var_path_src, browse_cmd=select_src_file)

# 操作行：一键处理 + 另存为 + 计算应纳税款 + 生成年度净值图表
make_action_row(panel1, row=3,
                action_text='一键处理',    action_cmd=process_and_preview,
                save_text='文件另存为',     save_cmd=save_book,
                save_var=var_save,
                vat_text='计算应纳税款',   vat_cmd=calc_vat,
                open_text='生成年度净值图表', open_cmd=compute_nav)

# 分隔线
separator1 = tk.Frame(main_frame, bg=COLORS['bg'], height=2)
separator1.grid(row=1, column=0, sticky=tk.EW)

# ────────────────────────────────────
# 面板 2：数据预览
# ────────────────────────────────────
panel2 = make_card_frame(main_frame, '数据预览', 2)

y_bar = tk.Scrollbar(panel2, orient=tk.VERTICAL)
y_bar.grid(row=0, column=2, sticky='ns', padx=(0, 4), pady=6)
x_bar = tk.Scrollbar(panel2, orient=tk.HORIZONTAL)
x_bar.grid(row=1, column=0, sticky='ew', padx=(8, 4), pady=(0, 6))

text_pad = tk.Text(
    panel2, height=14,
    bg=COLORS['text_bg'], fg=COLORS['text_fg'], font=FONTS['text'],
    xscrollcommand=x_bar.set, yscrollcommand=y_bar.set,
    relief=tk.FLAT, bd=0, padx=8, pady=6,
    selectbackground=COLORS['select_bg'], wrap=tk.NONE,
)
y_bar.config(command=text_pad.yview)
x_bar.config(command=text_pad.xview)
text_pad.grid(row=0, column=0, sticky=tk.NSEW, padx=(8, 4), pady=6)
panel2.grid_columnconfigure(0, weight=1)
panel2.grid_rowconfigure(0, weight=1)

text_pad.insert('1.0', '  欢迎使用券商数据处理程序')
text_pad.insert('insert', '\n')
text_pad.insert('insert', '  请按以下步骤操作：选择券商 → 导入格式文件 → 导入数据源 → 一键处理')
text_pad.insert('insert', '\n\n')
text_pad.insert('insert', '  ⚠️ 注意：所选待合并文件夹中仅能放置已经过本程序处理的文件')
text_pad.config(state=tk.DISABLED)

# 分隔线
separator2 = tk.Frame(main_frame, bg=COLORS['bg'], height=2)
separator2.grid(row=3, column=0, sticky=tk.EW)

# ────────────────────────────────────
# 面板 3：文件合并 & 分项归纳
# ────────────────────────────────────

panel3 = make_card_frame(main_frame, '文件合并', 4)

make_step_row(panel3, row=0, step_num='1', step_text='选择待合并文件夹',
              entry_var=var_path_merge, browse_cmd=select_merge_dir)

make_action_row(panel3, row=1,
                action_text='合并数据',     action_cmd=merge_files,
                save_text='另存为',          save_cmd=save_merged,
                save_var=var_save2,
                open_text='打开合并文件',    open_cmd=open_merged)

panel4 = make_card_frame(main_frame, '分项归纳', 5)

make_step_row(panel4, row=0, step_num='1', step_text='选择待归纳文件',
              entry_var=var_path_induct, browse_cmd=select_induct_file)

make_action_row(panel4, row=1,
                action_text='归纳数据',     action_cmd=induct_data,
                save_text='另存为',          save_cmd=save_inducted,
                save_var=var_save1,
                open_text='打开归纳文件',    open_cmd=open_inducted)


# ── 启动主事件循环 ──
root.mainloop()
